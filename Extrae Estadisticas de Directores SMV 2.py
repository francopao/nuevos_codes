# -*- coding: utf-8 -*-
"""
Extracción MASIVA y CONSOLIDADA de retribución de Directorio / Alta Gerencia /
Comités especiales desde los PDF 'SECCIONES B Y C' (Reporte de Cumplimiento del
Código BGC) descargados por 'descarga_memorias_smv.py'.

Recorre TODAS las subcarpetas de empresa dentro de RUTA_BASE, procesa cada PDF,
deduplica los documentos que la SMV subió dos veces, y genera UN solo Excel
consolidado con UNA HOJA POR EMPRESA (nombre de hoja = nombre canónico), donde
cada hoja apila, por ejercicio, las secciones relevantes:

    1) Resumen del Directorio (nº directores, independientes, tercio, sesiones,
       esquema y criterios de retribución, si aplica el modelo por sesión).
    2) Retribución como % de ingresos brutos: Directores (indep / no indep) y
       Alta Gerencia (fija / variable)  <-- tabla de la hoja ~46 (Pregunta III.19 a).
    3) Comités especiales (denominación, fecha, sesiones, % indep, participa JGA).
    4) MATRIZ DE PARTICIPACIÓN: filas = directores; columnas = Directorio + cada
       comité con su nº de sesiones. Es el insumo directo del bottom-up:
           sueldo = tarifa_sesion_dir * sesiones_directorio
                  + Σ (tarifa_sesion_comité * sesiones del comité donde participa)

Cada dato lleva una columna 'Pág.' con la hoja del PDF de donde salió, para que
puedas verificarlo rápido.

MEJORAS incorporadas:
  · MEJORA 1 — Equivalencia canónica de nombres de empresa contra el portafolio
    oficial. Todo nombre extraído se mapea a su nombre canónico antes de escribir;
    si no hay equivalencia razonable se conserva el original y se marca en
    'Incidencias' como 'sin equivalencia canónica'. El canónico se usa en (a) el
    título y nombre de hoja, (b) la cabecera de cada bloque de ejercicio y (c) el
    Índice. El agrupamiento por empresa también usa el canónico, de modo que las
    variantes (BBV/BBVA, ENGIE S.A./S.A.A., Nexa, Orygen, UNACEM…) colapsan en
    UNA sola hoja con todos sus ejercicios.
  · MEJORA 2 — Deduplicación robusta por tamaño en bytes + número de páginas.
    Si dos archivos coinciden EXACTAMENTE en bytes y páginas se consideran
    duplicados aunque difiera el nombre; el segundo NO se procesa y se registra
    en 'Incidencias' con ambos nombres, tamaño y páginas.
  · MEJORA 3 — Robustez de extracción:
      (A) PDFs sin capa de texto (escaneados) se detectan y se omiten, en vez de
          contaminar el libro con una hoja basura tomada de la ruta del archivo.
          El respaldo de nombre de empresa usa el NOMBRE DE ARCHIVO, nunca la ruta.
      (B) Detección de comités tolerante a encabezados 'COMITÉ N°/Nº/N 1' y, como
          red de seguridad, un segundo barrido que segmenta por filas
          'Denominación del Comité' cuando no hay encabezados 'COMITE N'.

ADVERTENCIA (no es bug, es el dato): el reporte solo informa la retribución como
% de ingresos brutos; NO da el monto por sesión. Y el criterio de pago varía por
empresa: la fórmula por sesión SOLO aplica donde el criterio fijo marque
"Por sesión de Directorio" (columna 'aplica_modelo_por_sesion').

Sin xlwings/win32com. Comentarios en español. Estructura por celdas (# %%) Spyder.
Dependencias:  pip install pdfplumber openpyxl pandas
"""

# %% ===========================  IMPORTS  ====================================
import os
import re
import glob
import unicodedata
from collections import defaultdict
from difflib import SequenceMatcher

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# %% ===========================  PARÁMETROS  =================================

RUTA_BASE = os.environ.get(
    "RUTA_BASE_OVERRIDE",
    r"C:\Users\usuario\OneDrive\Desktop\AFP INTEGRA\Docs privados\Otros\Remuneracion Directorio - Bottom Up\Remuneraciones",
)
SALIDA_XLSX = os.path.join(RUTA_BASE, "_consolidado_remuneracion_directorio.xlsx")

PATRON_ARCHIVO = "SECCIONES_B_Y_C"   # patrón en el nombre de archivo
UMBRAL_NOMBRE = 0.82                  # similitud mínima para emparejar nombres
UMBRAL_TEXTO = 200                   # MEJORA 3A: mínimo de chars para considerar el PDF legible


# %% ==================  PORTAFOLIO / NOMBRES CANÓNICOS  ======================
# MEJORA 1. Nombres oficiales del portafolio (respetando mayúsculas y tildes).
# A cada canónico se le asocian alias en minúsculas y sin tildes; el mapeo es
# por coincidencia de palabra completa sobre el nombre normalizado, de lo más
# específico (alias largo) a lo más corto, para que sea quirúrgico.
CANONICOS = {
    "Cerro Verde":       ["cerro verde"],
    "Pacasmayo":         ["pacasmayo"],
    "Aenza":             ["aenza"],
    "Ferreycorp":        ["ferreycorp"],
    "Engie":             ["engie"],
    "Volcan":            ["volcan"],
    "BBVA":              ["bbva", "bbv"],
    "Pluz Energia":      ["pluz"],
    "Puerto Chancay":    ["puerto chancay", "chancay", "cosco"],
    "Orygen":            ["orygen"],
    "Buenaventura":      ["buenaventura"],
    "Alicorp":           ["alicorp"],
    "UNACEM":            ["unacem"],
    "Credicorp":         ["credicorp"],
    "Nexa Perú":         ["nexa"],
    "InRetail":          ["inretail", "in retail"],
    "Auna":              ["auna"],
    "IFS":               ["intercorp financial", "ifs"],
    "Minsur":            ["minsur"],
    "Hudbay":            ["hudbay"],
    "Hunt Oil":          ["hunt oil", "hunt"],
    "Hermes":            ["hermes"],
    "Colegios Peruanos": ["colegios peruanos", "innova schools"],
    "Casa Andina":       ["casa andina"],
    "Inca Rail":         ["inca rail"],
    "Intursa":           ["intursa"],
    "Jockey Plaza":      ["jockey"],
    "Lima Expresa":      ["lima expresa", "lamsac"],
    "Orazen":            ["orazen"],
    "Primax":            ["primax"],
    "Rutas de Lima":     ["rutas de lima"],
    "Tecsup":            ["tecsup"],
    # 'Hochshild' tal cual lo escribe el portafolio. FOSSAL S.A.A. es el emisor
    # SMV del grupo Hochschild, por eso 'fossal' mapea aquí.
    "Hochshild":         ["hochshild", "hochschild", "fossal"],
}

# lista (alias, canónico) ordenada por longitud de alias desc (más específico primero)
_ALIAS_CANON = sorted(
    ((a, canon) for canon, lst in CANONICOS.items() for a in lst),
    key=lambda t: len(t[0]), reverse=True,
)


# %% ======================  UTILIDADES DE TEXTO  ============================

def norm(s):
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip()


def nlow(s):
    return norm(s).lower()


def _clave_emp(s):
    """Clave de comparación de empresa: sin tildes, minúsculas, signos -> espacio."""
    s = norm(s).lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def canoniza_empresa(raw):
    """MEJORA 1. Devuelve (nombre_canonico, encontrado_bool).
    Si no hay equivalencia razonable, devuelve (raw, False)."""
    clave = _clave_emp(raw)
    if not clave:
        return raw, False
    for alias, canon in _ALIAS_CANON:
        if re.search(r"\b" + re.escape(alias) + r"\b", clave):
            return canon, True
    return raw, False


FECHA = re.compile(r"\b(\d{2}/\d{2}/\d{4})\b")
# MEJORA 3B: título de comité tolerante. Se aplica con .match() sobre la clave
# normalizada (ver _clave_titulo_comite). Tolera 'COMITE 1', 'COMITE N 1',
# 'COMITE N° 1', 'COMITÉ Nº 1' (NFKD convierte 'º'->'o', de ahí la O? opcional)
# y 'Comité 6: ...' (texto posterior permitido al usar match en vez de fullmatch).
TITULO_COMITE = re.compile(r"^COMIT[EE]\s*(?:N\s*O?\s*)?(\d+)\b")
CARGOS_COMITE = ("presidente", "vicepresidente", "titular", "suplente",
                 "alterno", "secretario", "miembro")


def _clave_titulo_comite(c):
    """Normaliza una celda para probar si es un encabezado 'COMITE N':
    sin tildes, mayúsculas y todo signo (°, º, :, #, .) convertido a espacio."""
    s = unicodedata.normalize("NFKD", str(c))
    s = "".join(ch for ch in s if not unicodedata.combining(ch)).upper()
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _idx_titulo_comite(c):
    """Devuelve el índice (int) si la celda es un encabezado 'COMITE N', si no None."""
    m = TITULO_COMITE.match(_clave_titulo_comite(c))
    return int(m.group(1)) if m else None


def primer_numero(celdas):
    for c in celdas:
        m = re.search(r"-?\d+(?:\.\d+)?", norm(c))
        if m:
            return m.group(0)
    return None


def numeros_en(celdas):
    out = []
    for c in celdas:
        m = re.search(r"-?\d+(?:\.\d+)?", norm(c))
        if m:
            out.append(m.group(0))
    return out


def fila_si_no(celdas):
    idx_si = idx_no = idx_x = None
    for i, c in enumerate(celdas):
        n = nlow(c)
        if n in ("si", "sí") and idx_si is None:
            idx_si = i
        if n == "no" and idx_no is None:
            idx_no = i
        if n == "x" and idx_x is None:
            idx_x = i
    if idx_x is None or (idx_si is None and idx_no is None):
        return None
    if idx_si is not None and idx_no is not None:
        return "Sí" if abs(idx_x - idx_si) <= abs(idx_x - idx_no) else "No"
    return "Sí" if idx_si is not None else "No"


def parece_nombre(c):
    cl = nlow(c)
    if not c or FECHA.search(c) or len(c) < 5:
        return False
    if cl in CARGOS_COMITE:
        return False
    if any(k in cl for k in ["cargo", "fecha", "inicio", "termino", "facultades",
                             "ley general", "brindara", "comite", "sesiones",
                             "denominacion", "principales", "directores independientes",
                             "gerente general"]):
        return False
    return sum(1 for w in c.split() if w[:1].isupper()) >= 2


def similar(a, b):
    return SequenceMatcher(None, nlow(a), nlow(b)).ratio()


# %% ======================  EXTRACCIÓN POR PDF  =============================

def _flat_rows(pdf):
    """Filas de todas las tablas, cada una como (pagina_1based, [celdas norm])."""
    out = []
    for pi, pg in enumerate(pdf.pages):
        for tb in pg.extract_tables():
            for row in tb:
                out.append((pi + 1, [norm(c) for c in row]))
    return out


def _pagina_de(textos_pag, ancla):
    """Devuelve la 1ª página (1-based) cuyo texto contiene 'ancla' (normalizado)."""
    a = nlow(ancla)
    for i, t in enumerate(textos_pag):
        if a in nlow(t):
            return i + 1
    return None


def _nombre_desde_archivo(ruta):
    """MEJORA 3A. Respaldo de nombre de empresa a partir del NOMBRE DE ARCHIVO
    (nunca la ruta): quita el patrón, el año y la extensión, y deja palabras."""
    base = os.path.basename(ruta)
    base = re.sub(r"(?i)_?" + re.escape(PATRON_ARCHIVO) + r".*$", "", base)
    base = re.sub(r"(?i)\.pdf$", "", base)
    base = re.sub(r"\b\d{4}\b", " ", base)          # quita el año
    base = re.sub(r"[_\-]+", " ", base)
    return re.sub(r"\s+", " ", base).strip()


def extrae_pdf(ruta):
    with pdfplumber.open(ruta) as pdf:
        rows = _flat_rows(pdf)
        textos_pag = [(pg.extract_text() or "") for pg in pdf.pages]
    texto = "\n".join(textos_pag)
    p1 = textos_pag[0] if textos_pag else ""

    d = {"archivo": os.path.basename(ruta), "ruta": ruta, "pag": {}}

    # --- MEJORA 3A: PDF sin capa de texto (escaneado) -> ilegible, se omitirá ---
    if len(norm(texto)) < UMBRAL_TEXTO and not rows:
        d["pdf_ilegible"] = True
        return d

    # --- Empresa / ejercicio (pág. 1) ---
    emp = None
    lineas = p1.split("\n")
    for i, l in enumerate(lineas):
        nl = nlow(l)
        # MEJORA 3A: acepta 'Denominación:'/'Denominación social:'/'Razón social:'
        # tanto con el nombre en la MISMA línea como en la siguiente.
        if (nl.startswith("denominacion:") or nl.startswith("denominacion social:")
                or nl.startswith("razon social:")):
            resto = norm(l.split(":", 1)[1]) if ":" in l else ""
            if resto:
                emp = resto
            elif i + 1 < len(lineas):
                emp = norm(lineas[i + 1])
            break

    # MEJORA 3A: respaldo por NOMBRE DE ARCHIVO (nunca la ruta).
    emp_raw = emp or _nombre_desde_archivo(ruta)
    d["empresa"] = emp_raw

    # MEJORA 1: nombre canónico del portafolio.
    emp_canon, ok_canon = canoniza_empresa(emp_raw)
    d["empresa_canonica"] = emp_canon
    _inc_empresa = ([] if ok_canon else
                    [{"tipo": "sin equivalencia canónica",
                      "empresa": emp_raw, "archivo": d["archivo"]}])

    mej = re.search(r"Ejercicio:\s*(\d{4})", p1)
    d["ejercicio"] = int(mej.group(1)) if mej else None

    # --- Independientes (Principio 19) ---
    m_tot = re.search(r"De los\s+(\d+)\s+directores que conforman", texto)
    m_ind = re.search(r"(\d+)\s+son directores independientes", texto)
    d["n_directores_total"] = int(m_tot.group(1)) if m_tot else None
    d["n_independientes"] = int(m_ind.group(1)) if m_ind else None
    d["pag"]["independientes"] = _pagina_de(textos_pag, "son directores independientes")

    # ¿cumple el tercio? (arrastra encabezado Sí/No de la fila previa)
    d["cumple_tercio"] = None
    hdr_si = hdr_no = None
    for _, r in rows:
        c_si = next((i for i, c in enumerate(r) if nlow(c) in ("si", "sí")), None)
        c_no = next((i for i, c in enumerate(r) if nlow(c) == "no"), None)
        if c_si is not None and c_no is not None:
            hdr_si, hdr_no = c_si, c_no
        if "tercio del directorio" in nlow(" ".join(r)):
            sn = fila_si_no(r)
            if sn is None:
                idx_x = next((i for i, c in enumerate(r) if nlow(c) == "x"), None)
                if idx_x is not None and hdr_si is not None and hdr_no is not None:
                    sn = "Sí" if abs(idx_x - hdr_si) <= abs(idx_x - hdr_no) else "No"
            d["cumple_tercio"] = sn
            break

    # --- Retribución directores: esquema (f) y criterios (g) ---
    OPC = {
        "esquema": ["Fijo", "Variable", "Mixto (fijo + variable)"],
        "criterio_fijo": ["Por sesion de Directorio", "Por sesion de Comite",
                          "Por mes", "Por ano"],
        "criterio_variable": ["Por resultados del ejercicio",
                              "Por cumplimiento de objetivos"],
    }
    marcados = {k: [] for k in OPC}
    pag_esquema = None
    for pg, r in rows:
        if not any(nlow(c) == "x" for c in r):
            continue
        for grp, ops in OPC.items():
            for op in ops:
                if any(nlow(c) == nlow(op) for c in r):
                    marcados[grp].append(op)
                    if grp != "esquema":
                        pag_esquema = pg
    d["esquema_retribucion"] = ", ".join(dict.fromkeys(marcados["esquema"])) or None
    d["criterio_fijo"] = ", ".join(dict.fromkeys(marcados["criterio_fijo"])) or None
    d["criterio_variable"] = ", ".join(dict.fromkeys(marcados["criterio_variable"])) or None
    d["aplica_modelo_por_sesion"] = bool(
        re.search(r"por sesion de directorio", nlow(d["criterio_fijo"] or "")))
    d["paga_sesion_comite"] = bool(
        re.search(r"por sesion de comite", nlow(d["criterio_fijo"] or "")))
    d["pag"]["esquema"] = pag_esquema

    # --- Retribución directores como % de ingresos brutos (e) ---
    d["retrib_pct_no_indep"] = None
    d["retrib_pct_indep"] = None
    en_e = False
    for pg, r in rows:
        j = nlow(" ".join(r))
        if "retribuciones y de las bonificaciones" in j:
            en_e = True
            d["pag"]["retrib_dir"] = pg
            continue
        if en_e:
            if "directores (sin incluir" in j:
                d["retrib_pct_no_indep"] = primer_numero(
                    [c for c in r if not parece_nombre(c)])
            elif j.startswith("directores independientes"):
                d["retrib_pct_indep"] = primer_numero(r)
            elif "esquema de retribucion" in j or "f. precise" in j:
                en_e = False

    # --- Retribución Alta Gerencia (Pregunta III.19 a): Fija / Variable ---
    d["ger_fija_pct"] = None
    d["ger_variable_pct"] = None
    for pg, r in rows:
        if any("gerente general y plana gerencial" == nlow(c) for c in r):
            nums = numeros_en(r)
            if len(nums) >= 2:
                d["ger_fija_pct"], d["ger_variable_pct"] = nums[0], nums[1]
            elif len(nums) == 1:
                d["ger_fija_pct"] = nums[0]
            d["pag"]["retrib_gerencia"] = pg
            break

    # --- Sesiones del Directorio + asistencia por director (Principio 20) ---
    m_ses = re.search(r"N[uú]mero de sesiones realizadas\s+(\d+)", texto)
    d["n_sesiones_directorio"] = int(m_ses.group(1)) if m_ses else None

    asistencia = []
    en_tabla = False
    for pg, r in rows:
        j = nlow(" ".join(r))
        if "nombres y apellidos" in j and "convocadas" in j:
            en_tabla = True
            d["pag"]["sesiones"] = pg
            continue
        if en_tabla:
            nombre = next((c for c in r if parece_nombre(c)), None)
            nums = [c for c in r if re.fullmatch(r"\d+", norm(c))]
            if nombre and len(nums) >= 2:
                asistencia.append({"director": nombre,
                                   "sesiones_convocadas": int(nums[0]),
                                   "sesiones_asistio": int(nums[1])})
            elif r and not nombre and not nums and any(len(norm(c)) > 40 for c in r):
                en_tabla = False
    d["asistencia_directorio"] = asistencia
    if d["n_directores_total"] is None and asistencia:
        d["n_directores_total"] = len(asistencia)

    # --- Bloques COMITÉ N (Principio 21 'b.') ---
    comites = []
    cur = None
    for pg, r in rows:
        # MEJORA 3B: detección de título tolerante (COMITE N°/Nº/N 1, etc.)
        idx_tit = next((_idx_titulo_comite(c) for c in r
                        if _idx_titulo_comite(c) is not None), None)
        if idx_tit is not None:
            if cur:
                comites.append(cur)
            cur = {"comite_idx": idx_tit, "denominacion": None, "fecha_creacion": None,
                   "n_sesiones": None, "pct_indep": None, "participa_jga": None,
                   "pagina": pg, "miembros": []}
            continue
        if cur is None:
            continue
        j = nlow(" ".join(r))
        if "denominacion del comite" in j and cur["denominacion"] is None:
            v = [c for c in r if c and "denominacion" not in nlow(c)]
            cur["denominacion"] = v[0] if v else None
        elif "fecha de creacion" in j and cur["fecha_creacion"] is None:
            mf = next((FECHA.search(c) for c in r if FECHA.search(c)), None)
            if mf:
                cur["fecha_creacion"] = mf.group(1)
        elif "numero de sesiones realizadas" in j:
            cur["n_sesiones"] = primer_numero(
                [c for c in r if "numero de sesiones" not in nlow(c)])
        elif "directores independientes respec" in j:
            cur["pct_indep"] = primer_numero(
                [c for c in r if "independientes" not in nlow(c)])
        elif "participa en la jga" in j:
            cur["participa_jga"] = fila_si_no(r)
        else:
            nombre = next((c for c in r if parece_nombre(c)), None)
            if nombre and any(FECHA.search(c) for c in r):
                cargo = next((c for c in r if nlow(c) in CARGOS_COMITE), None)
                cur["miembros"].append({"nombre": nombre, "cargo_comite": cargo})
    if cur:
        comites.append(cur)
    comites = [c for c in comites if c["denominacion"]]

    # MEJORA 3B: red de seguridad. Si no se detectó NINGÚN comité por encabezado
    # 'COMITE N', se reintenta segmentando por filas 'Denominación del Comité'.
    if not comites:
        comites = _comites_por_denominacion(rows)

    d["comites"] = comites
    d["n_comites_especiales"] = len(d["comites"])

    # emparejar nombres de integrantes con el canónico del directorio
    d["incidencias_nombre"] = _canoniza_miembros(d) + _inc_empresa
    return d


def _comites_por_denominacion(rows):
    """MEJORA 3B. Respaldo: arma comités cuando NO hay encabezados 'COMITE N'.
    Abre un comité nuevo en cada fila 'Denominación del Comité' y le adjunta las
    filas de fecha / sesiones / % indep / JGA / miembros que vienen a continuación."""
    coms, cur, idx = [], None, 0
    for pg, r in rows:
        j = nlow(" ".join(r))
        if "denominacion del comite" in j:
            if cur:
                coms.append(cur)
            idx += 1
            v = [c for c in r if c and "denominacion" not in nlow(c)]
            cur = {"comite_idx": idx, "denominacion": v[0] if v else None,
                   "fecha_creacion": None, "n_sesiones": None, "pct_indep": None,
                   "participa_jga": None, "pagina": pg, "miembros": []}
            continue
        if cur is None:
            continue
        if "fecha de creacion" in j and cur["fecha_creacion"] is None:
            mf = next((FECHA.search(c) for c in r if FECHA.search(c)), None)
            if mf:
                cur["fecha_creacion"] = mf.group(1)
        elif "numero de sesiones realizadas" in j:
            cur["n_sesiones"] = primer_numero(
                [c for c in r if "numero de sesiones" not in nlow(c)])
        elif "directores independientes respec" in j:
            cur["pct_indep"] = primer_numero(
                [c for c in r if "independientes" not in nlow(c)])
        elif "participa en la jga" in j:
            cur["participa_jga"] = fila_si_no(r)
        else:
            nombre = next((c for c in r if parece_nombre(c)), None)
            if nombre and any(FECHA.search(c) for c in r):
                cargo = next((c for c in r if nlow(c) in CARGOS_COMITE), None)
                cur["miembros"].append({"nombre": nombre, "cargo_comite": cargo})
    if cur:
        coms.append(cur)
    return [c for c in coms if c["denominacion"]]


def _canoniza_miembros(d):
    canon = [a["director"] for a in d.get("asistencia_directorio", [])]
    inc = []
    for com in d.get("comites", []):
        for m in com["miembros"]:
            if not canon:
                m["director_canonico"] = m["nombre"]
                continue
            mejor = max(canon, key=lambda x: similar(x, m["nombre"]))
            if similar(mejor, m["nombre"]) >= UMBRAL_NOMBRE:
                m["director_canonico"] = mejor
            else:
                m["director_canonico"] = m["nombre"]
                inc.append({"empresa": d.get("empresa"), "ejercicio": d.get("ejercicio"),
                            "comite": com["denominacion"], "nombre_pdf": m["nombre"],
                            "mejor_candidato": mejor,
                            "score": round(similar(mejor, m["nombre"]), 2)})
    return inc


# %% ======================  ESTILOS / HELPERS EXCEL  =======================

F_TITULO = Font(bold=True, size=13, color="1F3864")
F_EJERC = Font(bold=True, size=11, color="FFFFFF")
F_SECCION = Font(bold=True, size=10, color="1F3864")
F_HEADER = Font(bold=True, size=9, color="FFFFFF")
F_NOTA = Font(italic=True, size=8, color="808080")
FILL_EJERC = PatternFill("solid", fgColor="1F3864")
FILL_HEADER = PatternFill("solid", fgColor="4472C4")
FILL_SECCION = PatternFill("solid", fgColor="D9E1F2")
FILL_APLICA = PatternFill("solid", fgColor="C6EFCE")
FILL_NOAPLICA = PatternFill("solid", fgColor="FFC7CE")
THIN = Side(style="thin", color="BFBFBF")
BORDE = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")


def _corto_comite(denom):
    """Nombre corto de comité para encabezado de la matriz."""
    s = norm(denom)
    s = re.sub(r"^Comite de\s+", "", s, flags=re.I)
    return (s[:22] + "…") if len(s) > 23 else s


def _alias_empresa(emp, usados):
    """Nombre de hoja reducido y único (<=31 chars). Recibe el nombre canónico."""
    e = norm(emp or "EMPRESA").upper()
    for suf in [" SOCIEDAD ANONIMA ABIERTA", " S.A.A.", " S.A.A", " S.A.C.",
                " S.A.", " S.A", " CORP.", " CORP", " LTD.", " LTD",
                " INC.", " INC", " PERU"]:
        e = e.replace(suf, "")
    e = re.sub(r"[\[\]\:\*\?\/\\]", "", e)
    e = re.sub(r"\s*-\s*.*$", "", e).strip()  # quita la parte tras un guion
    base = (e or "EMPRESA")[:28]
    nombre, k = base, 2
    while nombre in usados:
        nombre = f"{base[:28]}_{k}"
        k += 1
    usados.add(nombre)
    return nombre


def _bloque_tabla(ws, fila, titulo, headers, datos, anchos=None):
    """Escribe un bloque: título de sección + cabecera + filas. Devuelve la
    siguiente fila libre."""
    ncol = len(headers)
    ws.cell(fila, 1, titulo).font = F_SECCION
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=ncol)
    for c in range(1, ncol + 1):
        ws.cell(fila, c).fill = FILL_SECCION
    fila += 1
    for j, h in enumerate(headers, start=1):
        cel = ws.cell(fila, j, h)
        cel.font = F_HEADER
        cel.fill = FILL_HEADER
        cel.alignment = WRAP
        cel.border = BORDE
    fila += 1
    for row in datos:
        for j, v in enumerate(row, start=1):
            cel = ws.cell(fila, j, v)
            cel.border = BORDE
            cel.alignment = CENTER if j > 1 else WRAP
        fila += 1
    return fila + 1  # deja una fila en blanco


# %% ======================  ESCRITURA POR EMPRESA  =========================

def escribir_hoja_empresa(ws, registros):
    """registros: lista de dicts de UNA empresa (uno por ejercicio)."""
    regs = sorted(registros, key=lambda d: (d.get("ejercicio") or 0), reverse=True)
    # MEJORA 1: el título de la hoja es el nombre CANÓNICO.
    emp = next((r.get("empresa_canonica") for r in regs if r.get("empresa_canonica")),
               "EMPRESA")

    ws.cell(1, 1, emp).font = F_TITULO
    ws.freeze_panes = "A3"
    fila = 3

    for d in regs:
        ej = d.get("ejercicio")
        # MEJORA 1: el canónico encabeza cada bloque de ejercicio (celda empresa).
        ws.cell(fila, 1, f"{emp}   ·   EJERCICIO {ej}   ·   archivo: {d.get('archivo')}"
                ).font = F_EJERC
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=9)
        for c in range(1, 10):
            ws.cell(fila, c).fill = FILL_EJERC
        fila += 2

        pg = d.get("pag", {})
        # --- 1) Resumen del Directorio ---
        fila = _bloque_tabla(
            ws, fila, "1. Resumen del Directorio",
            ["N° Directores", "N° Independientes", "Cumple tercio",
             "Sesiones Directorio", "Esquema retribución", "Criterio fijo",
             "Criterio variable", "¿Aplica modelo por sesión?", "Pág."],
            [[d.get("n_directores_total"), d.get("n_independientes"),
              d.get("cumple_tercio"), d.get("n_sesiones_directorio"),
              d.get("esquema_retribucion"), d.get("criterio_fijo"),
              d.get("criterio_variable"),
              "SÍ" if d.get("aplica_modelo_por_sesion") else "NO",
              ", ".join(str(pg.get(k)) for k in ("independientes", "esquema", "sesiones")
                        if pg.get(k))]])
        # colorear la celda "aplica"
        cel_aplica = ws.cell(fila - 3, 8)
        cel_aplica.fill = FILL_APLICA if d.get("aplica_modelo_por_sesion") else FILL_NOAPLICA

        # --- 2) Retribución (% de ingresos brutos) ---
        fila = _bloque_tabla(
            ws, fila, "2. Retribución (% de ingresos brutos)",
            ["Directores no indep.", "Directores indep.",
             "Alta Gerencia – Fija", "Alta Gerencia – Variable",
             "Pág. directores", "Pág. gerencia"],
            [[d.get("retrib_pct_no_indep"), d.get("retrib_pct_indep"),
              d.get("ger_fija_pct"), d.get("ger_variable_pct"),
              pg.get("retrib_dir"), pg.get("retrib_gerencia")]])

        # --- 3) Comités especiales ---
        comites = d.get("comites", [])
        datos_com = [[c["denominacion"], c["fecha_creacion"], c["n_sesiones"],
                      c["pct_indep"], c["participa_jga"], len(c["miembros"]),
                      c["pagina"]] for c in comites]
        fila = _bloque_tabla(
            ws, fila, f"3. Comités especiales ({len(comites)})",
            ["Comité", "Fecha creación", "Sesiones", "% Indep.",
             "Participa JGA", "N° miembros", "Pág."],
            datos_com or [["(sin comités detectados)", "", "", "", "", "", ""]])

        # --- 4) Matriz de participación (bottom-up) ---
        asist = d.get("asistencia_directorio", [])
        asist_map = {a["director"]: a for a in asist}
        # miembros canónicos por comité
        sets_com = [{m.get("director_canonico", m["nombre"]) for m in c["miembros"]}
                    for c in comites]
        head = (["Director", "Ses. Dir. asistió", "Ses. Dir. convoc."]
                + [f"{_corto_comite(c['denominacion'])} ({c['n_sesiones']} ses)"
                   for c in comites]
                + ["Σ ses. comités"])
        filas_matriz = []
        # universo de directores = asistencia ∪ miembros de comités
        universo = [a["director"] for a in asist]
        for s in sets_com:
            for nom in s:
                if nom not in universo:
                    universo.append(nom)
        for nom in universo:
            a = asist_map.get(nom)
            row = [nom,
                   a["sesiones_asistio"] if a else "",
                   a["sesiones_convocadas"] if a else ""]
            total = 0
            for c, s in zip(comites, sets_com):
                if nom in s:
                    n = int(c["n_sesiones"]) if str(c["n_sesiones"]).isdigit() else None
                    row.append(n if n is not None else "✓")
                    total += (n or 0)
                else:
                    row.append("")
            row.append(total)
            filas_matriz.append(row)
        fila = _bloque_tabla(
            ws, fila,
            "4. Matriz de participación  (sesiones por órgano · insumo bottom-up)",
            head, filas_matriz or [["(sin datos)"] + [""] * (len(head) - 1)])

        ws.cell(fila, 1, "Nota: el reporte no informa monto por sesión ni asistencia "
                         "por miembro a cada comité; multiplique estas cantidades por "
                         "su tarifa por sesión.").font = F_NOTA
        fila += 3  # separación entre ejercicios

    # anchos de columna
    ws.column_dimensions["A"].width = 34
    for col in range(2, 14):
        ws.column_dimensions[get_column_letter(col)].width = 16


# %% ======================  EJECUCIÓN PRINCIPAL  ===========================

def localizar_pdfs(base):
    todos = glob.glob(os.path.join(base, "**", "*.pdf"), recursive=True)
    return sorted(p for p in todos
                  if PATRON_ARCHIVO.lower() in os.path.basename(p).lower())


def filtrar_duplicados_tam_pag(pdfs):
    """MEJORA 2. Descarta, ANTES de procesar, los PDF que coinciden EXACTAMENTE
    en (tamaño en bytes, nº de páginas) aunque difiera el nombre. Devuelve
    (pdfs_filtrados, incidencias). Sólo abre cada PDF para contar páginas."""
    vistos, quedan, inc = {}, [], []
    for ruta in pdfs:
        try:
            size = os.path.getsize(ruta)
            with pdfplumber.open(ruta) as pdf:
                npags = len(pdf.pages)
        except Exception:
            quedan.append(ruta)        # que el flujo normal lo intente/registre
            continue
        clave = (size, npags)
        if clave in vistos:
            inc.append({"tipo": "duplicado_tam_pag",
                        "archivo_conservado": os.path.basename(vistos[clave]),
                        "archivo_omitido": os.path.basename(ruta),
                        "tamano_bytes": size, "paginas": npags})
            print(f"  · duplicado (tam+pág) omitido: {os.path.basename(ruta)} "
                  f"== {os.path.basename(vistos[clave])} ({size} B, {npags} pág.)")
            continue
        vistos[clave] = ruta
        quedan.append(ruta)
    return quedan, inc


def main():
    pdfs = localizar_pdfs(RUTA_BASE)
    print(f"PDFs candidatos encontrados: {len(pdfs)}")

    # MEJORA 2: primera capa de dedup (tamaño + páginas), antes de procesar.
    pdfs, inc_tampag = filtrar_duplicados_tam_pag(pdfs)
    print(f"PDFs tras dedup por tamaño+páginas: {len(pdfs)}")

    registros, incidencias = [], list(inc_tampag)
    vistos = {}                      # (empresa_canon_norm, ejercicio) -> archivo (dedupe)
    por_empresa = defaultdict(list)  # empresa_canon_norm -> [registros]

    for ruta in pdfs:
        try:
            d = extrae_pdf(ruta)
        except Exception as e:
            incidencias.append({"tipo": "error_lectura",
                                "archivo": os.path.basename(ruta), "detalle": str(e)})
            print(f"  [!] error en {os.path.basename(ruta)}: {e}")
            continue

        # MEJORA 3A: PDF sin texto (escaneado) -> se omite y se registra.
        if d.get("pdf_ilegible"):
            incidencias.append({"tipo": "pdf_sin_texto", "archivo": d.get("archivo"),
                                "detalle": "sin capa de texto / escaneado; omitido"})
            print(f"  · PDF sin texto omitido: {d.get('archivo')}")
            continue

        # MEJORA 1: la clave de dedup y de agrupamiento usa el nombre CANÓNICO.
        emp_key = nlow(d.get("empresa_canonica") or d.get("empresa") or d.get("archivo"))
        clave = (emp_key, d.get("ejercicio"))
        if clave in vistos and d.get("ejercicio") is not None:
            incidencias.append({"tipo": "duplicado",
                                "empresa": d.get("empresa_canonica"),
                                "ejercicio": d.get("ejercicio"),
                                "archivo_omitido": d.get("archivo"),
                                "archivo_conservado": vistos[clave]})
            print(f"  · duplicado omitido: {d.get('archivo')}")
            continue
        vistos[clave] = d.get("archivo")

        incidencias.extend(d.pop("incidencias_nombre", []))
        registros.append(d)
        por_empresa[emp_key].append(d)
        print(f"  OK {d.get('empresa_canonica')} {d.get('ejercicio')} | "
              f"dir={d.get('n_directores_total')} indep={d.get('n_independientes')} "
              f"comites={d.get('n_comites_especiales')} "
              f"esquema={d.get('esquema_retribucion')} | ger_fija={d.get('ger_fija_pct')}")

    if not registros:
        print("No se extrajo ningún registro. Revisa RUTA_BASE y el patrón.")
        return

    # --- construir el libro ---
    wb = Workbook()
    wb.remove(wb.active)

    # hoja índice
    idx = wb.create_sheet("Índice")
    idx.cell(1, 1, "ÍNDICE DE EMPRESAS").font = F_TITULO
    idx.append([])
    for j, h in enumerate(["Empresa (canónica)", "Nombre extraído (PDF)", "Hoja",
                           "Ejercicios", "N° comités (último)"], start=1):
        c = idx.cell(3, j, h); c.font = F_HEADER; c.fill = FILL_HEADER; c.border = BORDE
    idx.freeze_panes = "A4"

    usados = set()
    fila_idx = 4
    for emp_key, regs in sorted(por_empresa.items()):
        # MEJORA 1: canónico para hoja e índice; conservamos el extraído como referencia.
        emp = next((r.get("empresa_canonica") for r in regs if r.get("empresa_canonica")),
                   emp_key)
        emp_extr = next((r.get("empresa") for r in regs if r.get("empresa")), "")
        hoja = _alias_empresa(emp, usados)
        ws = wb.create_sheet(hoja)
        escribir_hoja_empresa(ws, regs)
        ejs = ", ".join(str(e) for e in sorted(
            {r.get("ejercicio") for r in regs if r.get("ejercicio")}, reverse=True))
        ult = max(regs, key=lambda r: (r.get("ejercicio") or 0))
        for j, v in enumerate([emp, emp_extr, hoja, ejs,
                               ult.get("n_comites_especiales")], start=1):
            c = idx.cell(fila_idx, j, v); c.border = BORDE
        fila_idx += 1
    idx.column_dimensions["A"].width = 28
    idx.column_dimensions["B"].width = 40
    idx.column_dimensions["C"].width = 22
    idx.column_dimensions["D"].width = 18
    idx.column_dimensions["E"].width = 18

    # hoja incidencias
    inc = wb.create_sheet("Incidencias")
    if incidencias:
        cols = sorted({k for d in incidencias for k in d})
        for j, h in enumerate(cols, start=1):
            c = inc.cell(1, j, h); c.font = F_HEADER; c.fill = FILL_HEADER
        for i, d in enumerate(incidencias, start=2):
            for j, h in enumerate(cols, start=1):
                inc.cell(i, j, d.get(h))
    else:
        inc.cell(1, 1, "Sin incidencias")

    wb.save(SALIDA_XLSX)

    print("\n" + "=" * 70)
    print(f"Empresas (hojas)  : {len(por_empresa)}")
    print(f"Registros (años)  : {len(registros)}")
    print(f"Incidencias       : {len(incidencias)}")
    print(f"Excel generado    : {SALIDA_XLSX}")


# %% ======================  PUNTO DE ENTRADA  ==============================
if __name__ == "__main__":
    main()
