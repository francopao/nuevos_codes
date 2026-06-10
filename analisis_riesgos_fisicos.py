

# -*- coding: utf-8 -*-
"""
================================================================================
ANÁLISIS DE EXPOSICIÓN DE CARTERA A RIESGOS FÍSICOS (CENEPRED) — AFP INTEGRA
================================================================================
Qué hace este script
--------------------
1. Lee el Excel de PRODUCCIÓN MINERA (titular, unidad minera, depto/prov/distrito,
   unidad de medida y producción acumulada del año).
2. Convierte toda la producción a una unidad común: Tonelada Métrica Fina (TMF).
3. Geocodifica cada distrito a su UBIGEO usando un GeoJSON distrital del Perú.
4. Lee la hoja "CENEPRED" (score 1-5 por distrito para Mov. de Masa, Inundación
   y Sequías Severas).
5. Lee la hoja "Valorización de Instrumentos" y suma la exposición (S/) por
   empresa, aplicando una tabla de equivalencias de nombres (editable).
6. Distribuye la exposición de cada empresa entre los distritos donde produce
   (proporcional a su producción en TMF) y la cruza con el score CENEPRED para
   estimar la "plata en riesgo" por tipo de desastre y por ubicación.
7. Genera:
      - Un EXCEL de salida (una hoja por empresa + resúmenes).
      - Un HTML interactivo con el mapa coroplético del Perú.

Filosofía de blindaje
---------------------
- Las columnas se detectan por NOMBRE de encabezado (no por posición), con
  respaldo por letra de columna. Si la estructura cambia, el script aborta con
  un MENSAJE CLARO que indica qué buscaba, qué encontró y dónde mirar en el
  archivo para arreglarlo.
- El nombre del archivo es FLEXIBLE: se busca por patrón dentro de la carpeta.
- Toda la parametrización (rutas, equivalencias, correcciones de nombres de
  distritos, método de asignación) está en el bloque CONFIG de abajo, pensada
  para editarse a futuro sin tocar la lógica.
================================================================================
"""

import os
import re
import sys
import glob
import json
import unicodedata
from datetime import datetime

import pandas as pd
import numpy as np

# ==============================================================================
# 1) CONFIG  — EDITAR AQUÍ
# ==============================================================================

# Carpeta base del proyecto. Si la variable de entorno RF_BASE está definida,
# se usa esa (útil para pruebas); si no, se usa la ruta de AFP Integra.
BASE_DIR = os.environ.get(
    "RF_BASE",
    r"C:\Users\usuario\OneDrive\Desktop\AFP INTEGRA\ESG\Riesgos Fisicos\2026",
)

# Subcarpeta donde vive el Excel de producción minera (nombre flexible).
PROD_DIR = os.path.join(BASE_DIR, "Output")
# Patrones aceptados para el archivo de producción (el más reciente gana).
PROD_PATTERNS = ["Producci*Minera*.xls*", "*Producci*Minera*.xls*", "*Minera*.xls*"]

# Archivo de Riesgos Físicos (CENEPRED + Valorización). Nombre flexible y
# acepta .xlsx o .xlsm.
RISK_DIR = BASE_DIR
# El archivo real es "Riesgos Fisicos (Mapa Distrital) - 2026.xlsx" (el año varía).
# Apuntamos a él de forma específica para NO capturar los outputs del script
# (Riesgos_Fisicos_Cartera_*), que no contienen "Mapa Distrital".
RISK_PATTERNS = ["Riesgos Fisicos (Mapa Distrital) - *.xls*",
                 "Riesgos*Fisicos*Mapa*Distrital*.xls*",
                 "*Mapa*Distrital*.xls*"]

# Carpeta de salida (Excel + HTML).
OUT_DIR = os.path.join(BASE_DIR, "Inputs")

# GeoJSON distrital del Perú (con propiedades IDDIST/NOMBDEP/NOMBPROV/NOMBDIST).
# Se guarda/lee localmente; si no existe, se descarga de la URL.
GEOJSON_LOCAL = os.path.join(OUT_DIR, "peru_distritos.geojson")
GEOJSON_URL = ("https://raw.githubusercontent.com/juaneladio/peru-geojson/"
               "master/peru_distrital_simple.geojson")

# Nombres de hojas (con detección flexible si no se encuentran exactos).
SHEET_CENEPRED_HINTS = ["CENEPRED"]
SHEET_VALORIZ_HINTS = ["Valoriz", "Valorización", "Valorizacion", "Instrumentos"]

# --- Equivalencia de UNIDADES de medida a TMF (Tonelada Métrica Fina) ---------
#   TMF = Tonelada Métrica Fina            -> factor 1
#   Gramos finos (Grs. f.)                 -> TMF = gramos / 1_000_000
#   Kilogramos finos (Kg. f.)              -> TMF = kg / 1_000
# La clave se normaliza (sin espacios/puntos/acentos, en minúsculas).
UNIDAD_A_TMF = {
    "tmf": 1.0,
    "grsf": 1.0 / 1_000_000.0,   # gramos finos  -> TMF
    "kgf": 1.0 / 1_000.0,        # kilogramos finos -> TMF
}
# Etiquetas "bonitas" para mostrar la unidad original en el HTML.
UNIDAD_LABEL = {"tmf": "TMF", "grsf": "Grs. f.", "kgf": "Kg. f.",
                "credmiles": "Miles S/"}

# --- Equivalencia de NOMBRES de empresa (lado PRODUCCIÓN: titular legal) -------
# Editar/añadir libremente. Clave = nombre legal tal cual aparece en producción;
# Valor = nombre canónico (corto) con el que se reportará y se buscará exposición.
EQUIV_EMPRESAS_PRODUCCION = {
    "COMPAÑIA DE MINAS BUENAVENTURA S.A.A.": "Buenaventura",
    "SOCIEDAD MINERA EL BROCAL S.A.A.": "Buenaventura",
    "VOLCAN COMPAÑIA MINERA S.A.A.": "Volcan",
    "MINSUR S.A.": "Minsur",
    "NEXA RESOURCES PERU S.A.A.": "Nexa Resources Perú",
    "NEXA RESOURCES ATACOCHA S.A.A.": "Nexa Resources Perú",
    "NEXA RESOURCES EL PORVENIR S.A.C.": "Nexa Resources Perú",
    "NEXA RESOURCES CAJAMARQUILLA S.A.": "Nexa Resources Perú",
    "COMPAÑIA MINERA ARES S.A.C.": "Hochschild",
    "SOCIEDAD MINERA CERRO VERDE S.A.A.": "Cerro Verde",
    "MARCOBRE S.A.C.": "Minsur",
    "HUDBAY PERU S.A.C.": "Hudbay",
}

# --- Equivalencia de NOMBRES de empresa (lado EXPOSICIÓN: columna "Nombre (G&P)")
# Si la columna H trae variantes, mapéalas aquí al mismo nombre canónico.
# Los nombres que ya coinciden con el canónico no necesitan estar listados.
EQUIV_EMPRESAS_EXPOSICION = {
    "Nexa Resources": "Nexa Resources Perú",
}

# --- Correcciones manuales de geocodificación (cuando el nombre de distrito en
#     producción no calza con el GeoJSON). Clave = (DEP, PROV, DIST) normalizado;
#     valor = UBIGEO de 6 dígitos. Suele NO ser necesario gracias al respaldo
#     (departamento + distrito), pero queda aquí por si a futuro hace falta.
OVERRIDE_UBIGEO = {
    # ("ICA", "NASCA", "MARCONA"): "110304",
}

# --- Método de asignación de la exposición (S/) de cada empresa a sus distritos
#   "tmf_share"  : proporcional a la producción (TMF) de la empresa en cada distrito (recomendado)
#   "equal"      : repartida en partes iguales entre los distritos de la empresa
#   "units_share": proporcional al número de unidades mineras por distrito
METODO_ASIGNACION = "tmf_share"

# Empresas canónicas que son el universo de análisis (orden de reporte).
EMPRESAS_CANONICAS = ["Buenaventura", "Volcan", "Minsur", "Nexa Resources Perú",
                      "Hochschild", "Cerro Verde", "Hudbay"]

# Etiquetas de sector (se usan en el dashboard y el Excel).
SECTOR_MINERIA = "Minería"
SECTOR_FINANCIERO = "Financiero"

# Prefijos de los archivos que ESTE script genera. find_file los ignora siempre
# para no confundir un output (que puede caer en la misma carpeta) con un input.
EXCLUDE_OUTPUTS = ["Riesgos_Fisicos_Cartera", "Mapa_Riesgos_Fisicos"]

# ==============================================================================
# 1-bis) CONFIG DEL SECTOR FINANCIERO (BANCOS)
# ==============================================================================
# Analogía con minería: el "Total_Creditos_Suma" por distrito (en el archivo ya
# procesado B-2358-resumen-*.xlsx) cumple el mismo rol que la PRODUCCIÓN (TMF) en
# minería: es el PESO con el que se reparte geográficamente la exposición de
# cartera de cada banco entre los distritos donde coloca créditos.
#
# Interruptor maestro: pon en False para volver al análisis solo-minería.
INCLUIR_FINANCIERO = True

# Carpeta donde vive el resumen de bancos. Por defecto la raíz "AFP INTEGRA"
# (un nivel arriba de ...\ESG\Riesgos Fisicos\2026). Override con la variable de
# entorno RF_FIN si lo tienes en otro lado.
FIN_DIR = os.environ.get(
    "RF_FIN",
    os.path.dirname(os.path.dirname(os.path.dirname(BASE_DIR))),
)
# Patrones de nombre aceptados (el más reciente gana). El archivo ESPERADO es el
# RESUMEN ya procesado por tu script Financials (columnas Empresa, Departamento,
# Provincia, Distrito, Total_Creditos_Suma), NO el crudo B-2358 de la SBS.
FIN_PATTERNS = ["B-2358-resumen-*.xlsx", "B-2358-resumen-*.xls*",
                "*resumen*B-2358*.xls*", "*resumen*creditos*.xls*"]

# Nombres ESPERADOS de columna en el resumen de bancos (se buscan de forma robusta
# a mayúsculas/minúsculas y tildes; si no se encuentran, el script dice cuál falta).
FIN_COLS = {
    "empresa": ["Empresa", "Banco", "Entidad"],
    "depto":   ["Departamento", "Depto", "Region"],
    "prov":    ["Provincia"],
    "dist":    ["Distrito"],
    "monto":   ["Total_Creditos_Suma", "Total Creditos Suma", "Total Creditos",
                "Total_Creditos", "Creditos", "Monto"],
}

# --- Equivalencia de NOMBRES de banco (lado CRÉDITOS: columna A del resumen) ---
# Clave = nombre tal cual aparece en la columna "Empresa" del resumen de bancos;
# Valor = nombre canónico del grupo en el portafolio. El emparejamiento es robusto
# a mayúsculas/minúsculas, tildes y espacios extra (se normaliza antes de comparar).
# Si la SBS cambia un nombre en el futuro, el script te avisará por consola qué
# bancos del archivo NO están mapeados aquí para que los agregues.
EQUIV_EMPRESAS_FINANCIERO = {
    "BANCO DE CREDITO": "Credicorp",
    "INTERBANK": "Intercorp Perú",
    "BBVA": "BBVA",
}

# Bancos canónicos que forman el universo financiero de análisis (orden de reporte).
EMPRESAS_FIN_CANONICAS = ["Credicorp", "Intercorp Perú", "BBVA"]

# --- Equivalencia de NOMBRES (lado EXPOSICIÓN: columna H "Nombre (G&P)") --------
# Igual que en minería, en "Valorización de Instrumentos" un mismo grupo económico
# puede aparecer bajo varias razones sociales. Aquí decides QUÉ líneas de la col H
# se SUMAN a cada banco canónico. Es un juicio de negocio: revísalo.
#   OJO con los homónimos que NO son del grupo peruano y por eso quedan FUERA por
#   defecto (descoméntalos solo si decides incluirlos):
#     - "BBVA Bancomer"                  -> BBVA México (no BBVA Perú)
#     - "Banco de Crédito e Inversiones" -> BCI de Chile (no BCP/Credicorp)
# El script imprime una auditoría con lo que sumó a cada grupo y los nombres de la
# col H que parecen bancarios pero quedaron sin mapear, para que ajustes esta lista.
EQUIV_EXPOSICION_FINANCIERO = {
    "Credicorp": "Credicorp",
    "Banco de Crédito del Perú": "Credicorp",
    "Intercorp Perú": "Intercorp Perú",
    "Interbank": "Intercorp Perú",
    "BBVA": "BBVA",
    # "BBVA Bancomer": "BBVA",                        # ← BBVA México (revisar)
    # "Banco de Crédito e Inversiones": "Credicorp",  # ← BCI Chile (revisar)
}

# Pistas para detectar, en la auditoría, nombres de col H que "parecen banco" y
# podrías querer mapear (solo para avisarte; no afecta el cálculo).
PISTAS_BANCO = ["BANCO", "BANK", "CREDICORP", "INTERCORP", "INTERBANK", "BBVA",
                "SCOTIA", "FINANCI", "CREDITO", "BCP"]

HAZARDS = ["Movimientos de Masa", "Inundación", "Sequías Severas"]
HAZARD_KEY = {"Movimientos de Masa": "mm", "Inundación": "inu", "Sequías Severas": "seq"}

# Paleta corporativa AFP Integra
AFP_AZUL = "#1E2E6E"
AFP_CYAN = "#00AECB"
AFP_AMAR = "#E3E829"


# ==============================================================================
# 2) UTILIDADES
# ==============================================================================

def log(msg, level="INFO"):
    print(f"[{level}] {msg}")


def fatal(msg):
    print("\n" + "=" * 78)
    print("ERROR — el proceso se detuvo. Revisa el detalle de abajo:")
    print("=" * 78)
    print(msg)
    print("=" * 78)
    sys.exit(1)


def norm(s):
    """Normaliza texto: mayúsculas, sin acentos, solo [A-Z0-9 ], colapsa espacios."""
    if s is None:
        return ""
    s = str(s).strip().upper()
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def norm_unidad(s):
    """Normaliza la unidad de medida: 'Grs. f.' -> 'grsf', 'Kg.f.' -> 'kgf', 'TMF' -> 'tmf'."""
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]", "", s)


def find_file(folder, patterns, descripcion, exclude=None):
    if not os.path.isdir(folder):
        fatal(f"No existe la carpeta esperada para {descripcion}:\n  {folder}\n"
              f"-> Verifica la ruta en CONFIG (BASE_DIR) o crea/renombra la carpeta.")
    hits = []
    for pat in patterns:
        hits += glob.glob(os.path.join(folder, pat))
    # excluir temporales de Excel (~$...)
    hits = [h for h in sorted(set(hits)) if not os.path.basename(h).startswith("~$")]
    # excluir los archivos que el propio script genera (evita autoconfusión cuando
    # un output cae en la misma carpeta y comparte prefijo de nombre con un input)
    excl = (exclude or []) + EXCLUDE_OUTPUTS
    if excl:
        hits = [h for h in hits
                if not any(norm(s) in norm(os.path.basename(h)) for s in excl)]
    if not hits:
        fatal(f"No se encontró ningún archivo de {descripcion} en:\n  {folder}\n"
              f"Patrones buscados: {patterns}\n"
              f"-> Confirma que el archivo esté en esa carpeta. El NOMBRE puede "
              f"variar, pero debe contener las palabras del patrón.\n"
              f"   (Se ignoran los archivos de salida del propio script: {EXCLUDE_OUTPUTS}.)")
    hits.sort(key=os.path.getmtime, reverse=True)
    if len(hits) > 1:
        log(f"Se encontraron {len(hits)} candidatos para {descripcion}; se usa el "
            f"más reciente: {os.path.basename(hits[0])}", "WARN")
    return hits[0]


def pick_sheet(xls_path, hints, descripcion):
    xl = pd.ExcelFile(xls_path)
    sheets = xl.sheet_names
    for h in hints:
        for s in sheets:
            if norm(h) in norm(s):
                return s
    fatal(f"No se encontró la hoja de {descripcion} en {os.path.basename(xls_path)}.\n"
          f"Hojas disponibles: {sheets}\n"
          f"Pistas buscadas: {hints}\n"
          f"-> Renombra la hoja o agrega su nombre a los *_HINTS en CONFIG.")


def detect_header_row(df_raw, required_tokens, max_scan=20):
    """Busca la fila de encabezados: la que contiene más de los tokens esperados."""
    req = [norm(t) for t in required_tokens]
    best_i, best_score = None, -1
    for i in range(min(max_scan, len(df_raw))):
        cells = [norm(x) for x in df_raw.iloc[i].tolist()]
        score = sum(1 for t in req if any(t in c or c in t for c in cells if c))
        if score > best_score:
            best_i, best_score = i, score
    return best_i, best_score


def resolve_columns(columns, spec):
    """
    spec: {logico: ([substrings de encabezado], letra_respaldo_opcional)}
    Devuelve {logico: nombre_real_de_columna}. Lanza fatal si falta alguno.
    """
    norm_cols = {col: norm(col) for col in columns}
    out = {}
    faltan = []
    for logico, (subs, letra) in spec.items():
        found = None
        for col, ncol in norm_cols.items():
            if any(norm(sub) in ncol for sub in subs):
                found = col
                break
        if found is None and letra:  # respaldo por letra de columna
            idx = column_letter_to_index(letra)
            if idx < len(columns):
                found = columns[idx]
                log(f"Columna '{logico}' no se halló por nombre; se usa respaldo por "
                    f"letra {letra} -> '{found}'.", "WARN")
        if found is None:
            faltan.append((logico, subs, letra))
        else:
            out[logico] = found
    if faltan:
        det = "\n".join(f"   - {lg}: buscaba encabezados {subs} (respaldo letra {lt})"
                        for lg, subs, lt in faltan)
        fatal("No se pudieron ubicar estas columnas en el Excel:\n" + det +
              f"\n\nColumnas detectadas en el archivo:\n   {list(columns)}\n"
              "-> Abre el Excel y revisa que esos encabezados existan (o ajusta los "
              "'subs' en la función correspondiente del script).")
    return out


def column_letter_to_index(letter):
    letter = letter.upper()
    idx = 0
    for ch in letter:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx - 1


# ==============================================================================
# 3) CARGA DEL GEOJSON Y CONSTRUCCIÓN DEL ÍNDICE NOMBRE -> UBIGEO
# ==============================================================================

def load_geojson():
    if os.path.isfile(GEOJSON_LOCAL):
        log(f"GeoJSON local: {GEOJSON_LOCAL}")
        with open(GEOJSON_LOCAL, "r", encoding="utf-8") as f:
            return json.load(f)
    # descargar
    log("No hay GeoJSON local; intentando descargar...", "WARN")
    try:
        import urllib.request
        os.makedirs(os.path.dirname(GEOJSON_LOCAL), exist_ok=True)
        with urllib.request.urlopen(GEOJSON_URL, timeout=90) as r:
            data = json.loads(r.read().decode("utf-8"))
        with open(GEOJSON_LOCAL, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False)
        log(f"GeoJSON descargado y guardado en {GEOJSON_LOCAL}")
        return data
    except Exception as e:
        fatal(f"No se pudo cargar ni descargar el GeoJSON distrital.\n"
              f"Detalle: {e}\n"
              f"-> Descarga manualmente este archivo y guárdalo como:\n"
              f"   {GEOJSON_LOCAL}\n"
              f"   Fuente: {GEOJSON_URL}")


def build_geo_index(geojson):
    """Devuelve:
       name2u : {(dep,prov,dist)->ubigeo}, dep_dist2u : {(dep,dist)->[ubigeos]},
       u2name : {ubigeo->(dep,prov,dist)} con nombres ORIGINALES."""
    name2u, dep_dist2u, u2name = {}, {}, {}
    for f in geojson["features"]:
        p = f.get("properties", {})
        u = p.get("IDDIST")
        if not u:
            continue
        u = str(u).strip().zfill(6)
        dep, prov, dist = p.get("NOMBDEP"), p.get("NOMBPROV"), p.get("NOMBDIST")
        name2u[(norm(dep), norm(prov), norm(dist))] = u
        dep_dist2u.setdefault((norm(dep), norm(dist)), []).append(u)
        u2name[u] = (dep, prov, dist)
    return name2u, dep_dist2u, u2name


def geocode(dep, prov, dist, idx):
    """Resuelve UBIGEO: 1) override 2) (dep,prov,dist) 3) (dep,dist) único.
       Devuelve (ubigeo, metodo) o (None, motivo)."""
    name2u, dep_dist2u, _ = idx
    k3 = (norm(dep), norm(prov), norm(dist))
    if k3 in OVERRIDE_UBIGEO:
        return OVERRIDE_UBIGEO[k3], "override"
    if k3 in name2u:
        return name2u[k3], "exacto"
    if norm(prov) in ("", "-") or norm(dist) in ("", "-"):
        return None, "sin_distrito"
    cands = dep_dist2u.get((norm(dep), norm(dist)), [])
    if len(cands) == 1:
        return cands[0], "dep+distrito"
    if len(cands) > 1:
        return None, "ambiguo"
    return None, "no_encontrado"


# ==============================================================================
# 4) LECTURA DE PRODUCCIÓN MINERA
# ==============================================================================

def load_produccion(path):
    xl = pd.ExcelFile(path)
    sheet = xl.sheet_names[0]
    if len(xl.sheet_names) > 1:
        # preferir una hoja que parezca de "informacion general / produccion"
        for s in xl.sheet_names:
            if norm("INFORMACION") in norm(s) or norm("PRODUCC") in norm(s):
                sheet = s
                break
    raw = pd.read_excel(path, sheet_name=sheet, header=None)
    hdr_row, score = detect_header_row(
        raw, ["MINERAL", "UNIDAD DE MEDIDA", "TITULAR", "DEPARTAMENTO",
              "DISTRITO", "ACUM"])
    if score < 4:
        fatal(f"No se reconoció la fila de encabezados en la hoja '{sheet}' de "
              f"producción (coincidencias={score}).\n"
              f"-> Confirma que existan encabezados como MINERAL, UNIDAD DE MEDIDA, "
              f"TITULAR, DEPARTAMENTO, PROVINCIA, DISTRITO y un acumulado anual (ACUM...).")
    df = pd.read_excel(path, sheet_name=sheet, header=hdr_row)
    df = df.dropna(axis=1, how="all")

    spec = {
        "mineral":  (["MINERAL"], "A"),
        "unidad":   (["UNIDAD DE MEDIDA", "UNIDAD"], "B"),
        "titular":  (["TITULAR"], "F"),
        "unidad_minera": (["UNIDAD MINERA"], "G"),
        "depto":    (["DEPARTAMENTO"], "H"),
        "prov":     (["PROVINCIA"], "I"),
        "dist":     (["DISTRITO"], "J"),
        "acum":     (["ACUM"], "W"),
    }
    cols = resolve_columns(df.columns, spec)
    df = df.rename(columns={v: k for k, v in cols.items()})

    # filtrar pies de página / filas sin titular o sin departamento
    df = df[df["titular"].notna() & df["depto"].notna()].copy()
    # acumulado numérico; si viene vacío, intentar sumar meses (respaldo)
    df["acum"] = pd.to_numeric(df["acum"], errors="coerce")
    if df["acum"].isna().all():
        meses = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO",
                 "AGOSTO", "SETIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
        mcols = [c for c in df.columns if norm(c) in [norm(m) for m in meses]]
        if mcols:
            df["acum"] = df[mcols].apply(pd.to_numeric, errors="coerce").sum(axis=1)
            log("Acumulado anual reconstruido sumando columnas mensuales.", "WARN")
        else:
            fatal("La columna de producción acumulada está vacía y no se hallaron "
                  "columnas mensuales para reconstruirla. Revisa el Excel de producción.")
    df["acum"] = df["acum"].fillna(0.0)

    # unidad -> factor TMF
    df["unidad_norm"] = df["unidad"].apply(norm_unidad)
    desconocidas = sorted(set(df.loc[~df["unidad_norm"].isin(UNIDAD_A_TMF), "unidad"]
                              .dropna().astype(str)))
    if desconocidas:
        log(f"Unidades de medida no reconocidas (se excluyen del TMF): {desconocidas}\n"
            f"      -> Si corresponden, agrégalas a UNIDAD_A_TMF en CONFIG.", "WARN")
    df["factor"] = df["unidad_norm"].map(UNIDAD_A_TMF)
    df["tmf"] = df["acum"] * df["factor"]

    # empresa canónica (solo universo de interés)
    eq_norm = {norm(k): v for k, v in EQUIV_EMPRESAS_PRODUCCION.items()}
    df["empresa"] = df["titular"].apply(lambda t: eq_norm.get(norm(t)))
    df_univ = df[df["empresa"].notna()].copy()
    log(f"Producción: {len(df)} filas válidas; {len(df_univ)} de las empresas del "
        f"universo ({df_univ['empresa'].nunique()} empresas).")
    return df_univ, sheet


# ==============================================================================
# 4-bis) LECTURA DEL RESUMEN DE BANCOS (CRÉDITOS POR DISTRITO)
# ==============================================================================

def _resolver_columnas_por_nombre(df_cols, fin_cols, descripcion):
    """Empareja los nombres ESPERADOS (robustos a mayúsculas/tildes) con las
    columnas reales del DataFrame. Devuelve {clave: nombre_real}. Si falta alguna,
    aborta con un mensaje claro de qué columna falta y cómo arreglarlo."""
    norm_real = {norm(c): c for c in df_cols}
    resol = {}
    faltan = []
    for clave, alias in fin_cols.items():
        encontrado = None
        for a in alias:
            if norm(a) in norm_real:
                encontrado = norm_real[norm(a)]
                break
        if encontrado is None:
            faltan.append((clave, alias))
        else:
            resol[clave] = encontrado
    if faltan:
        detalle = "\n".join(
            f"      · '{clave}': se buscó alguno de {alias}" for clave, alias in faltan)
        fatal(f"En {descripcion} no se encontraron estas columnas:\n{detalle}\n"
              f"   Columnas presentes en el archivo: {list(df_cols)}\n"
              f"-> Ajusta los nombres esperados en FIN_COLS dentro de CONFIG, o "
              f"corrige el encabezado del archivo. (No importan mayúsculas ni tildes.)")
    return resol


def load_financiero(path):
    """Lee el resumen de bancos (Empresa, Departamento, Provincia, Distrito,
    Total_Creditos_Suma) y lo devuelve con la MISMA forma que la producción minera
    para reutilizar todo el pipeline: el monto de créditos pasa a la columna 'tmf'
    (el peso de reparto), y se etiqueta como pseudo-mineral 'Créditos'."""
    xl = pd.ExcelFile(path)
    # preferir una hoja que parezca de resumen
    sheet = xl.sheet_names[0]
    for s in xl.sheet_names:
        if norm("RESUMEN") in norm(s):
            sheet = s
            break

    raw = pd.read_excel(path, sheet_name=sheet, header=None)
    hdr_row, _ = detect_header_row(
        raw, ["Empresa", "Departamento", "Provincia", "Distrito"], max_scan=15)
    df = pd.read_excel(path, sheet_name=sheet, header=hdr_row)
    df = df.dropna(axis=1, how="all")

    cols = _resolver_columnas_por_nombre(df.columns, FIN_COLS,
                                         "el resumen de bancos (créditos)")
    df = df.rename(columns={
        cols["empresa"]: "titular", cols["depto"]: "depto",
        cols["prov"]: "prov", cols["dist"]: "dist", cols["monto"]: "acum"})

    df = df[df["titular"].notna() & df["depto"].notna()].copy()
    df["acum"] = pd.to_numeric(df["acum"], errors="coerce").fillna(0.0)

    # --- empresa canónica (robusto a mayúsculas/tildes/espacios) ---
    eq_norm = {norm(k): v for k, v in EQUIV_EMPRESAS_FINANCIERO.items()}
    df["empresa"] = df["titular"].apply(lambda t: eq_norm.get(norm(t)))

    # Aviso de bancos del archivo que NO están mapeados (para que decidas incluirlos)
    no_map = sorted(set(str(t).strip() for t in df.loc[df["empresa"].isna(), "titular"]
                        .dropna()))
    if no_map:
        log(f"Bancos en el resumen NO mapeados al portafolio (se omiten). Si alguno "
            f"debe incluirse, agrégalo a EQUIV_EMPRESAS_FINANCIERO en CONFIG:", "WARN")
        for b in no_map:
            log(f"      · '{b}'", "WARN")

    df_univ = df[df["empresa"].notna()].copy()
    if df_univ.empty:
        log("Ningún banco del resumen coincide con EQUIV_EMPRESAS_FINANCIERO; el "
            "sector financiero quedará vacío. Revisa los nombres en CONFIG.", "WARN")

    # dar forma de 'producción' para reutilizar build_model sin cambios
    df_univ["mineral"] = "Créditos"
    df_univ["unidad"] = "Miles S/"
    df_univ["unidad_minera"] = None
    df_univ["unidad_norm"] = "credmiles"     # etiqueta propia (no entra a UNIDAD_A_TMF)
    df_univ["factor"] = 1.0
    df_univ["tmf"] = df_univ["acum"]          # el crédito ES el peso de reparto

    log(f"Financiero: {len(df)} filas leídas; {len(df_univ)} de bancos del universo "
        f"({df_univ['empresa'].nunique()} bancos).")
    return df_univ[["mineral", "unidad", "titular", "unidad_minera", "depto", "prov",
                    "dist", "acum", "unidad_norm", "factor", "tmf", "empresa"]], sheet


# ==============================================================================
# 5) LECTURA DE CENEPRED
# ==============================================================================

def load_cenepred(path, sheet):
    raw = pd.read_excel(path, sheet_name=sheet, header=None)
    # fila de ubigeos = la que tiene más celdas tipo código de 6 dígitos
    ubi_row = None
    for i in range(min(10, len(raw))):
        vals = raw.iloc[i].tolist()[1:]
        cnt = sum(1 for v in vals if re.fullmatch(r"\d{5,6}", str(v).strip().split(".")[0] if v is not None else ""))
        if cnt > 30:
            ubi_row = i
            break
    if ubi_row is None:
        fatal(f"No se encontró la fila de UBIGEOS en la hoja CENEPRED.\n"
              f"-> Debe haber una fila con códigos de distrito de 6 dígitos "
              f"(p.ej. 010201) a lo largo de las columnas.")
    ubigeos = []
    for v in raw.iloc[ubi_row].tolist()[1:]:
        if v is None or str(v).strip() == "":
            ubigeos.append(None)
        else:
            ubigeos.append(str(v).strip().split(".")[0].zfill(6))
    # filas de peligros: las siguientes con etiqueta en la 1a columna
    scores = {}  # ubigeo -> {mm,inu,seq}
    hazard_rows = {}
    for i in range(ubi_row + 1, len(raw)):
        label = raw.iloc[i, 0]
        if label is None or str(label).strip() == "":
            continue
        nl = norm(label)
        for hz in HAZARDS:
            if norm(hz) in nl or nl in norm(hz):
                hazard_rows[hz] = i
    faltan = [h for h in HAZARDS if h not in hazard_rows]
    if faltan:
        fatal(f"En CENEPRED no se hallaron las filas de peligro: {faltan}\n"
              f"-> La primera columna debe contener: {HAZARDS}.")
    for hz, ri in hazard_rows.items():
        rowvals = raw.iloc[ri].tolist()[1:]
        for j, u in enumerate(ubigeos):
            if u is None or j >= len(rowvals):
                continue
            val = rowvals[j]
            try:
                sc = int(float(val))
            except (TypeError, ValueError):
                sc = None
            scores.setdefault(u, {})[HAZARD_KEY[hz]] = sc
    log(f"CENEPRED: {len([u for u in ubigeos if u])} distritos con score.")
    return scores


# ==============================================================================
# 6) LECTURA DE EXPOSICIÓN (VALORIZACIÓN DE INSTRUMENTOS)
# ==============================================================================

def load_exposicion(path, sheet):
    raw = pd.read_excel(path, sheet_name=sheet, header=None)
    hdr_row, score = detect_header_row(raw, ["Nombre (G&P)", "VALOR_VECTOR_LOCAL"],
                                       max_scan=10)
    if score < 1:
        fatal(f"No se reconoció la fila de encabezados en '{sheet}'.\n"
              f"-> Debe existir una columna 'Nombre (G&P)' (H) y "
              f"'VALOR_VECTOR_LOCAL' (M).")
    df = pd.read_excel(path, sheet_name=sheet, header=hdr_row)
    df = df.dropna(how="all")
    spec = {
        "nombre": (["Nombre (G&P)", "Nombre G P", "Nombre"], "H"),
        "valor":  (["VALOR_VECTOR_LOCAL", "VALOR VECTOR LOCAL"], "M"),
    }
    cols = resolve_columns(df.columns, spec)
    df = df.rename(columns={cols["nombre"]: "nombre", cols["valor"]: "valor"})
    df["nombre"] = df["nombre"].astype(str).str.strip()
    df["valor"] = pd.to_numeric(df["valor"], errors="coerce").fillna(0.0)

    eq_ex_norm = {norm(k): v for k, v in EQUIV_EMPRESAS_EXPOSICION.items()}
    canon_norm = {norm(c): c for c in EMPRESAS_CANONICAS}

    def to_canon(n):
        nn = norm(n)
        if nn in eq_ex_norm:
            return eq_ex_norm[nn]
        if nn in canon_norm:
            return canon_norm[nn]
        return None

    df["empresa"] = df["nombre"].apply(to_canon)
    exp = (df.dropna(subset=["empresa"]).groupby("empresa")["valor"].sum().to_dict())
    for c in EMPRESAS_CANONICAS:
        exp.setdefault(c, 0.0)
    log(f"Exposición: S/ {sum(exp.values()):,.0f} en {sum(1 for v in exp.values() if v>0)} empresas.")
    return exp


# ==============================================================================
# 7) CONSTRUCCIÓN DEL MODELO (producción por distrito, asignación, riesgo)
# ==============================================================================

def build_model(df_prod, exp, scores, geo_idx, universo=None, sector=SECTOR_MINERIA):
    if universo is None:
        universo = EMPRESAS_CANONICAS
    _, _, u2name = geo_idx
    # geocodificar cada fila
    geocodes, motivos = [], []
    for _, r in df_prod.iterrows():
        u, m = geocode(r["depto"], r["prov"], r["dist"], geo_idx)
        geocodes.append(u)
        motivos.append(m)
    df = df_prod.copy()
    df["ubigeo"] = geocodes
    df["geo_motivo"] = motivos

    no_map = df[df["ubigeo"].isna()].copy()
    if len(no_map):
        resumen = (no_map.groupby(["depto", "prov", "dist", "geo_motivo"])
                   .size().reset_index(name="filas"))
        log(f"{len(no_map)} filas de producción NO geocodificadas "
            f"({no_map['tmf'].sum():.2f} TMF). Detalle por motivo:", "WARN")
        for _, rr in resumen.iterrows():
            log(f"      {rr['depto']} / {rr['prov']} / {rr['dist']}  "
                f"[{rr['geo_motivo']}]  filas={rr['filas']}", "WARN")

    dfm = df[df["ubigeo"].notna()].copy()

    # ---- producción por (empresa, ubigeo) en TMF + detalle por mineral ----
    prod_emp_dist = (dfm.groupby(["empresa", "ubigeo"])["tmf"].sum()
                     .reset_index())
    # unidades mineras por (empresa, ubigeo)
    um = (dfm.groupby(["empresa", "ubigeo"])["unidad_minera"]
          .apply(lambda s: sorted(set(str(x) for x in s.dropna()))).to_dict())
    # detalle por mineral (en unidad original) por (empresa, ubigeo)
    detalle = {}
    for (e, u, mineral, uni_norm), g in dfm.groupby(
            ["empresa", "ubigeo", "mineral", "unidad_norm"]):
        detalle.setdefault((e, u), []).append({
            "mineral": str(mineral),
            "unidad": UNIDAD_LABEL.get(uni_norm, uni_norm),
            "cantidad_original": float(g["acum"].sum()),
            "tmf": float(g["tmf"].sum()),
        })

    # ---- asignación de exposición de cada empresa a sus distritos ----
    df_assign = prod_emp_dist.copy()
    df_assign["exposicion"] = 0.0
    for e in universo:
        sub = df_assign[df_assign["empresa"] == e]
        E = exp.get(e, 0.0)
        if sub.empty or E == 0:
            continue
        if METODO_ASIGNACION == "tmf_share" and sub["tmf"].sum() > 0:
            w = sub["tmf"] / sub["tmf"].sum()
        elif METODO_ASIGNACION == "units_share":
            nun = sub["ubigeo"].map(lambda u: max(1, len(um.get((e, u), []))))
            w = nun / nun.sum()
        else:  # equal o tmf=0
            w = pd.Series(1.0 / len(sub), index=sub.index)
        df_assign.loc[sub.index, "exposicion"] = (E * w).values

    # ---- consolidado por distrito (exposición) ----
    dist_rows = {}
    for _, r in df_assign.iterrows():
        u = r["ubigeo"]
        d = dist_rows.setdefault(u, {"ubigeo": u, "tmf": 0.0, "exposicion": 0.0,
                                     "empresas": {}})
        d["tmf"] += r["tmf"]
        d["exposicion"] += r["exposicion"]
        d["empresas"][r["empresa"]] = d["empresas"].get(r["empresa"], 0.0) + r["exposicion"]

    # ---- desglose por mineral (unidad original) agregado por ubigeo (para tooltip) ----
    min_by_u = {}
    for (e, u), items in detalle.items():
        agg = min_by_u.setdefault(u, {})
        for it in items:
            key = (it["mineral"], it["unidad"])
            a = agg.setdefault(key, {"cantidad_original": 0.0, "tmf": 0.0})
            a["cantidad_original"] += it["cantidad_original"]
            a["tmf"] += it["tmf"]

    # Universo de distritos a publicar: TODOS los que tienen score CENEPRED (y existen
    # en el geojson, para que pinten en el mapa) UNIDOS a los que tienen exposición.
    # Esto permite mostrar el panorama completo de riesgo físico del país y superponer
    # la plata invertida solo donde corresponde.
    ubigeos_score = [u for u in scores if u in u2name]
    todos_u = set(dist_rows) | set(ubigeos_score)
    score_no_geo = [u for u in scores if u not in u2name]
    if score_no_geo:
        log(f"{len(score_no_geo)} distrito(s) con score CENEPRED no existen en el "
            f"geojson y no se pintarán en el mapa: {', '.join(sorted(score_no_geo))}",
            "WARN")

    districts = []
    for u in todos_u:
        dep, prov, dist = u2name.get(u, (None, None, None))
        d = dist_rows.get(u, {"tmf": 0.0, "exposicion": 0.0, "empresas": {}})
        sc = scores.get(u, {})
        mm, inu, seq = sc.get("mm"), sc.get("inu"), sc.get("seq")
        avail = [x for x in (mm, inu, seq) if x is not None]
        comb = round(sum(avail) / len(avail), 2) if avail else None
        minerales = [
            {"mineral": k[0], "unidad": k[1], "sector": sector,
             "cantidad_original": round(v["cantidad_original"], 3),
             "tmf": round(v["tmf"], 4)}
            for k, v in min_by_u.get(u, {}).items()
        ]
        expo = round(d["exposicion"], 2)
        rec = {
            "ubigeo": u, "dep": dep, "prov": prov, "dist": dist,
            "tmf": round(d["tmf"], 4), "exposicion": expo,
            "exp_sec": {sector: expo} if expo > 0 else {},
            "has_exp": expo > 0,
            "has_score": comb is not None,
            "score_mm": mm, "score_inu": inu, "score_seq": seq, "score_comb": comb,
            "mar_mm": round(expo * mm, 2) if mm else 0.0,
            "mar_inu": round(expo * inu, 2) if inu else 0.0,
            "mar_seq": round(expo * seq, 2) if seq else 0.0,
            "empresas": {k: round(v, 2) for k, v in d["empresas"].items() if v > 0},
            "minerales": minerales,
        }
        districts.append(rec)
    return districts, df_assign, um, detalle, no_map, dfm


# ==============================================================================
# 7-bis) EXPOSICIÓN FINANCIERA + AUDITORÍA + FUSIÓN DE SECTORES
# ==============================================================================

def load_exposicion_financiero(path, sheet):
    """Exposición de cartera a bancos desde 'Valorización de Instrumentos'
    (col H = nombre, col M = VALOR_VECTOR_LOCAL), usando el mapeo editable
    EQUIV_EXPOSICION_FINANCIERO. Imprime una auditoría de qué nombres de la col H
    se sumaron a cada grupo y cuáles 'parecen banco' pero quedaron sin mapear."""
    raw = pd.read_excel(path, sheet_name=sheet, header=None)
    hdr_row, _ = detect_header_row(raw, ["Nombre (G&P)", "VALOR_VECTOR_LOCAL"],
                                   max_scan=10)
    df = pd.read_excel(path, sheet_name=sheet, header=hdr_row).dropna(how="all")
    spec = {"nombre": (["Nombre (G&P)", "Nombre G P", "Nombre"], "H"),
            "valor":  (["VALOR_VECTOR_LOCAL", "VALOR VECTOR LOCAL"], "M")}
    cols = resolve_columns(df.columns, spec)
    df = df.rename(columns={cols["nombre"]: "nombre", cols["valor"]: "valor"})
    df["nombre"] = df["nombre"].astype(str).str.strip()
    df["valor"] = pd.to_numeric(df["valor"], errors="coerce").fillna(0.0)

    eq_norm = {norm(k): v for k, v in EQUIV_EXPOSICION_FINANCIERO.items()}
    df["empresa"] = df["nombre"].apply(lambda n: eq_norm.get(norm(n)))

    exp = (df.dropna(subset=["empresa"]).groupby("empresa")["valor"].sum().to_dict())
    for c in EMPRESAS_FIN_CANONICAS:
        exp.setdefault(c, 0.0)

    # ---- auditoría ----
    log(f"Exposición financiera: S/ {sum(exp.values()):,.0f} en "
        f"{sum(1 for v in exp.values() if v > 0)} bancos.")
    detalle_grupo = (df.dropna(subset=["empresa"])
                     .groupby(["empresa", "nombre"])["valor"].sum())
    for canon in EMPRESAS_FIN_CANONICAS:
        if canon in detalle_grupo.index.get_level_values(0):
            partes = detalle_grupo.loc[canon].sort_values(ascending=False)
            comp = "; ".join(f"{nm}: S/ {v:,.0f}" for nm, v in partes.items())
            log(f"      [{canon}] = {comp}")
    # nombres que parecen banco y NO se mapearon
    no_map = df[df["empresa"].isna()]
    cand = (no_map[no_map["nombre"].apply(lambda n: any(p in norm(n) for p in PISTAS_BANCO))]
            .groupby("nombre")["valor"].sum().sort_values(ascending=False))
    if len(cand):
        log("      Nombres en col H que parecen banco y NO están mapeados "
            "(revisa EQUIV_EXPOSICION_FINANCIERO si alguno corresponde):", "WARN")
        for nm, v in cand.items():
            log(f"        · '{nm}'  (S/ {v:,.0f})", "WARN")
    return exp


def merge_sectores(listas_districts):
    """Fusiona varias listas de 'districts' (una por sector) en una sola por ubigeo.
    Suma exposición y plata-en-riesgo, combina empresas y concatena el desglose.
    Los scores CENEPRED son idénticos por ubigeo entre sectores."""
    out = {}
    for districts in listas_districts:
        for d in districts:
            u = d["ubigeo"]
            if u not in out:
                out[u] = {
                    "ubigeo": u, "dep": d["dep"], "prov": d["prov"], "dist": d["dist"],
                    "tmf": 0.0, "exposicion": 0.0, "exp_sec": {},
                    "has_exp": False, "has_score": d["has_score"],
                    "score_mm": d["score_mm"], "score_inu": d["score_inu"],
                    "score_seq": d["score_seq"], "score_comb": d["score_comb"],
                    "mar_mm": 0.0, "mar_inu": 0.0, "mar_seq": 0.0,
                    "empresas": {}, "minerales": [],
                }
            o = out[u]
            o["tmf"] += d.get("tmf", 0.0)
            o["exposicion"] = round(o["exposicion"] + d["exposicion"], 2)
            for s, v in d.get("exp_sec", {}).items():
                o["exp_sec"][s] = round(o["exp_sec"].get(s, 0.0) + v, 2)
            o["has_exp"] = o["has_exp"] or d["has_exp"]
            for hz in ("mar_mm", "mar_inu", "mar_seq"):
                o[hz] = round(o[hz] + d.get(hz, 0.0), 2)
            for k, v in d.get("empresas", {}).items():
                o["empresas"][k] = round(o["empresas"].get(k, 0.0) + v, 2)
            o["minerales"].extend(d.get("minerales", []))
    return list(out.values())


# ==============================================================================
# 8) EXCEL DE SALIDA
# ==============================================================================

def export_excel(out_path, df_assign, exp, scores, geo_idx, um, detalle,
                 districts, no_map, universo=None, emp_sector=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if universo is None:
        universo = EMPRESAS_CANONICAS
    if emp_sector is None:
        emp_sector = {e: SECTOR_MINERIA for e in universo}

    def es_fin(e):
        return emp_sector.get(e) == SECTOR_FINANCIERO

    def peso_lbl(e, largo=False):
        if es_fin(e):
            return "Créditos colocados (miles S/)" if largo else "Créditos (miles S/)"
        return "Producción total (TMF)" if largo else "Producción (TMF)"

    _, _, u2name = geo_idx
    wb = Workbook()
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor=AFP_AZUL.replace("#", ""))
    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    title_font = Font(name="Calibri", bold=True, color=AFP_AZUL.replace("#", ""), size=14)
    sub_font = Font(name="Calibri", italic=True, color="595959", size=9)

    def style_header(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

    total_exp = sum(exp.values())

    # ---------- Hoja RESUMEN ----------
    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = "Exposición de cartera a riesgos físicos — Resumen por empresa"
    ws["A1"].font = title_font
    ws["A2"] = (f"Generado: {datetime.now():%Y-%m-%d %H:%M}  ·  Minería en TMF "
                f"(Tonelada Métrica Fina) · Banca en miles de S/ colocados")
    ws["A2"].font = sub_font
    headers = ["Empresa (canónica)", "Sector", "Producción / Créditos", "N° distritos",
               "Exposición cartera (S/)", "% de la exposición total"]
    ws.append([])
    ws.append(headers)
    hrow = ws.max_row
    style_header(ws, hrow, len(headers))
    prod_emp = df_assign.groupby("empresa")["tmf"].sum().to_dict()
    ndist_emp = df_assign[df_assign["tmf"] > 0].groupby("empresa")["ubigeo"].nunique().to_dict()
    for e in universo:
        ws.append([e, emp_sector.get(e, ""), round(prod_emp.get(e, 0.0), 2),
                   int(ndist_emp.get(e, 0)), round(exp.get(e, 0.0), 2), None])
        r = ws.max_row
    # total + porcentajes con fórmulas (A=Empresa B=Sector C=Prod D=NDist E=Exp F=%)
    first = hrow + 1
    last = ws.max_row
    ws.append(["TOTAL", "", f"=SUM(C{first}:C{last})", f"=SUM(D{first}:D{last})",
               f"=SUM(E{first}:E{last})", None])
    trow = ws.max_row
    for i, e in enumerate(universo):
        rr = first + i
        ws.cell(rr, 6).value = f"=IF($E${trow}=0,0,E{rr}/$E${trow})"
        ws.cell(rr, 6).number_format = "0.0%"
    ws.cell(trow, 6).value = f"=IF($E${trow}=0,0,SUM(E{first}:E{last})/$E${trow})"
    ws.cell(trow, 6).number_format = "0.0%"
    for c in range(1, 7):
        ws.cell(trow, c).font = Font(name="Calibri", bold=True)
    for rr in range(first, trow + 1):
        ws.cell(rr, 3).number_format = "#,##0.00"
        ws.cell(rr, 5).number_format = "#,##0"
    widths = [26, 13, 22, 13, 22, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{hrow + 1}"

    # ---------- Hoja por EMPRESA ----------
    for e in universo:
        sub = df_assign[(df_assign["empresa"] == e)].copy()
        ws = wb.create_sheet(_safe_sheet_name(e))
        et = "Colocaciones por distrito y exposición asignada" if es_fin(e) \
            else "Producción por distrito y exposición asignada"
        ws["A1"] = f"{e} — {et}"
        ws["A1"].font = title_font
        tot_lbl = "Créditos totales" if es_fin(e) else "Producción total"
        tot_uni = "miles S/" if es_fin(e) else "TMF"
        ws["A2"] = (f"{tot_lbl}: {sub['tmf'].sum():,.2f} {tot_uni}   ·   "
                    f"Exposición en cartera: S/ {exp.get(e,0):,.0f}   ·   "
                    f"Asignación: {METODO_ASIGNACION}")
        ws["A2"].font = sub_font
        if sub.empty:
            ws["A4"] = "Sin registros para esta empresa en el archivo."
            continue
        col_peso = "Créditos (miles S/)" if es_fin(e) else "Producción (TMF)"
        col_pct = "% del banco" if es_fin(e) else "% prod. empresa"
        col_um = "Oficinas/Agencias" if es_fin(e) else "Unidades mineras"
        headers = ["Departamento", "Provincia", "Distrito", "Ubigeo",
                   col_um, col_peso, col_pct,
                   "Exposición asignada (S/)", "Score Mov. Masa", "Score Inundación",
                   "Score Sequías", "Plata en riesgo (S/, comb.)"]
        ws.append([]); ws.append(headers)
        hrow = ws.max_row
        style_header(ws, hrow, len(headers))
        sub = sub.sort_values("tmf", ascending=False)
        first = hrow + 1
        for _, r in sub.iterrows():
            u = r["ubigeo"]
            dep, prov, dist = u2name.get(u, ("", "", ""))
            sc = scores.get(u, {})
            uni = ", ".join(um.get((e, u), [])) or "—"
            avail = [sc.get(k) for k in ("mm", "inu", "seq") if sc.get(k) is not None]
            comb = sum(avail) / len(avail) if avail else 0
            mar = r["exposicion"] * comb
            ws.append([dep, prov, dist, u, uni, round(r["tmf"], 4), None,
                       round(r["exposicion"], 2), sc.get("mm"), sc.get("inu"),
                       sc.get("seq"), round(mar, 2)])
        last = ws.max_row
        tot_tmf_cell = f"=SUM(F{first}:F{last})"
        ws.append(["TOTAL", "", "", "", "", tot_tmf_cell, "", f"=SUM(H{first}:H{last})",
                   "", "", "", f"=SUM(L{first}:L{last})"])
        trow = ws.max_row
        for i in range(first, last + 1):
            ws.cell(i, 7).value = f"=IF($F${trow}=0,0,F{i}/$F${trow})"
            ws.cell(i, 7).number_format = "0.0%"
            ws.cell(i, 6).number_format = "#,##0.0000"
            ws.cell(i, 8).number_format = "#,##0"
            ws.cell(i, 12).number_format = "#,##0"
        for c in (1, 6, 8, 12):
            ws.cell(trow, c).font = Font(bold=True)
        ws.cell(trow, 6).number_format = "#,##0.0000"
        ws.cell(trow, 8).number_format = "#,##0"
        ws.cell(trow, 12).number_format = "#,##0"
        for col, w in zip("ABCDEFGHIJKL",
                          [16, 16, 18, 9, 34, 16, 13, 20, 13, 13, 12, 20]):
            ws.column_dimensions[col].width = w
        ws.freeze_panes = f"A{first}"

        # detalle por mineral / colocaciones (debajo)
        ws.append([]); ws.append([])
        det_t = "Detalle de colocaciones (miles S/)" if es_fin(e) \
            else "Detalle por mineral (unidad original)"
        ws.append([det_t])
        ws.cell(ws.max_row, 1).font = Font(bold=True, color=AFP_AZUL.replace("#", ""))
        dh = ["Distrito", "Ubigeo", "Mineral", "Unidad", "Cantidad (orig.)", "TMF equiv."]
        ws.append(dh)
        style_header(ws, ws.max_row, len(dh))
        for u in sub["ubigeo"]:
            dep, prov, dist = u2name.get(u, ("", "", ""))
            for it in detalle.get((e, u), []):
                ws.append([dist, u, it["mineral"], it["unidad"],
                           round(it["cantidad_original"], 4), round(it["tmf"], 6)])
                ws.cell(ws.max_row, 5).number_format = "#,##0.0000"
                ws.cell(ws.max_row, 6).number_format = "#,##0.000000"

    # ---------- Hoja EXPOSICIÓN Y RIESGO (por distrito) ----------
    ws = wb.create_sheet("Exposición y Riesgo")
    ws["A1"] = "Exposición y plata en riesgo por distrito (todas las empresas)"
    ws["A1"].font = title_font
    ws["A2"] = ("Plata en riesgo = exposición asignada × score CENEPRED (1-5). "
                "Score combinado = promedio de los disponibles.")
    ws["A2"].font = sub_font
    headers = ["Departamento", "Provincia", "Distrito", "Ubigeo",
               "Exposición (S/)", "Score Mov. Masa", "Score Inundación",
               "Score Sequías", "Score comb.", "Riesgo Mov. Masa (S/)",
               "Riesgo Inundación (S/)", "Riesgo Sequías (S/)"]
    ws.append([]); ws.append(headers)
    style_header(ws, ws.max_row, len(headers))
    for d in sorted(districts, key=lambda x: x["exposicion"], reverse=True):
        ws.append([d["dep"], d["prov"], d["dist"], d["ubigeo"], d["exposicion"],
                   d["score_mm"], d["score_inu"], d["score_seq"], d["score_comb"],
                   d["mar_mm"], d["mar_inu"], d["mar_seq"]])
        for c in (5, 10, 11, 12):
            ws.cell(ws.max_row, c).number_format = "#,##0"
    for col, w in zip("ABCDEFGHIJKL", [16, 16, 18, 9, 18, 13, 13, 12, 11, 18, 18, 18]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"

    # ---------- Hoja EQUIVALENCIAS ----------
    ws = wb.create_sheet("Equivalencias")
    ws["A1"] = "Equivalencias de unidades y de nombres de empresa"
    ws["A1"].font = title_font
    ws.append([]); ws.append(["Unidades de medida (todo se lleva a TMF)"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, color=AFP_AZUL.replace("#", ""))
    ws.append(["Abreviatura", "Significado", "Inglés", "Uso típico", "Factor a TMF"])
    style_header(ws, ws.max_row, 5)
    uni_tbl = [
        ["TMF", "Tonelada Métrica Fina", "Metric Fine Ton (MFT)",
         "Cobre, zinc, plomo, molibdeno, hierro, estaño", 1],
        ["Grs. f.", "Gramos finos", "Fine grams", "Oro", "1 / 1,000,000"],
        ["Kg. f.", "Kilogramos finos", "Fine kilograms", "Plata", "1 / 1,000"],
    ]
    for row in uni_tbl:
        ws.append(row)
    ws.append([]); ws.append(["Equivalencia de empresas (producción → canónico)"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, color=AFP_AZUL.replace("#", ""))
    ws.append(["Nombre en producción (TITULAR)", "Nombre canónico"])
    style_header(ws, ws.max_row, 2)
    for k, v in EQUIV_EMPRESAS_PRODUCCION.items():
        ws.append([k, v])
    ws.append([]); ws.append(["Equivalencia de empresas (exposición → canónico)"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, color=AFP_AZUL.replace("#", ""))
    ws.append(["Nombre en Valorización (col. H)", "Nombre canónico"])
    style_header(ws, ws.max_row, 2)
    for k, v in EQUIV_EMPRESAS_EXPOSICION.items():
        ws.append([k, v])
    for col, w in zip("ABCDE", [40, 26, 24, 44, 14]):
        ws.column_dimensions[col].width = w

    # ---------- Hoja NO MAPEADOS ----------
    ws = wb.create_sheet("No mapeados")
    ws["A1"] = "Filas de producción sin ubigeo asignado (excluidas del mapa)"
    ws["A1"].font = title_font
    ws.append([]); ws.append(["Empresa", "Departamento", "Provincia", "Distrito",
                              "Motivo", "Producción (TMF)"])
    style_header(ws, ws.max_row, 6)
    if len(no_map):
        agg = (no_map.groupby(["empresa", "depto", "prov", "dist", "geo_motivo"])["tmf"]
               .sum().reset_index())
        for _, r in agg.iterrows():
            ws.append([r["empresa"], r["depto"], r["prov"], r["dist"],
                       r["geo_motivo"], round(r["tmf"], 4)])
            ws.cell(ws.max_row, 6).number_format = "#,##0.0000"
    else:
        ws.append(["(Todas las filas fueron geocodificadas correctamente)"])
    for col, w in zip("ABCDEF", [22, 18, 18, 20, 16, 16]):
        ws.column_dimensions[col].width = w

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return out_path


def _safe_sheet_name(name):
    n = re.sub(r"[\[\]\:\*\?\/\\]", " ", name).strip()[:31]
    return n


# ==============================================================================
# 9) HTML INTERACTIVO
# ==============================================================================

def export_html(out_path, geojson, districts, exp, scores, geo_idx, emp_sector=None):
    _, _, u2name = geo_idx
    if emp_sector is None:
        emp_sector = {e: SECTOR_MINERIA for e in exp}
    exposicion_sector = {}
    for e, v in exp.items():
        s = emp_sector.get(e, SECTOR_MINERIA)
        exposicion_sector[s] = round(exposicion_sector.get(s, 0.0) + v, 2)
    payload = {
        "generado": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "districts": districts,
        "exposicion_empresa": {k: round(v, 2) for k, v in exp.items()},
        "emp_sector": emp_sector,
        "exposicion_sector": exposicion_sector,
        "sectores": [SECTOR_MINERIA, SECTOR_FINANCIERO],
        "metodo": METODO_ASIGNACION,
        "afp": {"azul": AFP_AZUL, "cyan": AFP_CYAN, "amar": AFP_AMAR},
    }
    geojson_min = {"type": "FeatureCollection", "features": [
        {"type": "Feature",
         "properties": {"IDDIST": str(f["properties"].get("IDDIST", "")).zfill(6),
                        "NOMBDIST": f["properties"].get("NOMBDIST"),
                        "NOMBPROV": f["properties"].get("NOMBPROV"),
                        "NOMBDEP": f["properties"].get("NOMBDEP")},
         "geometry": f["geometry"]}
        for f in geojson["features"] if f.get("geometry")]}

    html = _HTML_TEMPLATE.replace("__PAYLOAD__", json.dumps(payload, ensure_ascii=False))
    html = html.replace("__GEOJSON__", json.dumps(geojson_min, ensure_ascii=False))
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    return out_path


_HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="es"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>Exposición de Cartera a Riesgos Físicos — AFP Integra</title>
<script src="https://cdn.plot.ly/plotly-2.35.2.min.js" charset="utf-8"></script>
<style>
 :root{--azul:#1E2E6E;--cyan:#00AECB;--amar:#E3E829;--gris:#f4f5f8;--txt:#1f2533;}
 *{box-sizing:border-box;font-family:Calibri,'Segoe UI',Arial,sans-serif;}
 body{margin:0;background:#fff;color:var(--txt);}
 header{background:var(--azul);color:#fff;padding:18px 26px;}
 header h1{margin:0;font-size:22px;font-weight:700;}
 header p{margin:4px 0 0;font-size:13px;opacity:.85;}
 .wrap{max-width:1340px;margin:0 auto;padding:18px 22px 60px;}
 .kpis{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin:18px 0;}
 .kpi{background:var(--gris);border-left:5px solid var(--cyan);border-radius:8px;padding:14px 16px;}
 .kpi .v{font-size:22px;font-weight:700;color:var(--azul);}
 .kpi .l{font-size:12px;color:#5a6275;margin-top:4px;line-height:1.25;}
 .controls{display:flex;flex-wrap:wrap;gap:14px;align-items:flex-end;margin:10px 0 16px;}
 .controls label{font-size:12px;color:#5a6275;display:block;margin-bottom:4px;font-weight:600;}
 select{padding:8px 10px;border:1px solid #cfd4e0;border-radius:7px;font-size:14px;min-width:230px;background:#fff;}
 .grid{display:grid;grid-template-columns:1.35fr 1fr;gap:18px;}
 .card{border:1px solid #e6e8ef;border-radius:10px;padding:14px 16px;margin-top:18px;}
 .card h3{margin:0 0 8px;color:var(--azul);font-size:16px;}
 .note{font-size:12px;color:#6b7280;line-height:1.5;}
 table{border-collapse:collapse;width:100%;font-size:13px;}
 th,td{padding:7px 9px;border-bottom:1px solid #eef0f5;text-align:left;}
 th{background:var(--azul);color:#fff;position:sticky;top:0;cursor:pointer;}
 td.num,th.num{text-align:right;font-variant-numeric:tabular-nums;}
 tr:hover td{background:#f7f9ff;}
 .scroll{max-height:360px;overflow:auto;border:1px solid #eef0f5;border-radius:8px;}
 .pill{display:inline-block;background:var(--gris);border-radius:20px;padding:2px 10px;font-size:11px;margin:2px 4px 2px 0;}
 .legend{font-size:12px;color:#5a6275;margin-top:6px;}
 .tag{font-weight:700;color:var(--azul);}
 details{margin-top:10px;}
 summary{cursor:pointer;font-weight:600;color:var(--azul);}
 @media(max-width:980px){.kpis{grid-template-columns:repeat(2,1fr);}.grid{grid-template-columns:1fr;}}
</style></head>
<body>
<header>
  <h1>Exposición de cartera a riesgos físicos (CENEPRED)</h1>
  <p id="sub">AFP Integra · ESG · Riesgos Físicos</p>
</header>
<div class="wrap">
  <div class="kpis" id="kpis"></div>

  <div class="controls">
    <div><label>Sector:</label>
      <select id="sector">
        <option value="todos" selected>Todos los sectores</option>
        <option value="Minería">Minería</option>
        <option value="Financiero">Financiero (bancos)</option>
      </select>
    </div>
    <div><label>Colorear el mapa por:</label>
      <select id="metric">
        <optgroup label="Riesgo físico (score CENEPRED 1-5)">
          <option value="score_mm">Movimientos de Masa</option>
          <option value="score_inu">Inundación</option>
          <option value="score_seq">Sequías Severas</option>
          <option value="score_comb">Score combinado (promedio)</option>
        </optgroup>
        <optgroup label="Cartera">
          <option value="exposicion">Exposición asignada (S/)</option>
        </optgroup>
        <optgroup label="Plata en riesgo (exposición × score)">
          <option value="mar_mm">Riesgo · Movimientos de Masa (S/)</option>
          <option value="mar_inu">Riesgo · Inundación (S/)</option>
          <option value="mar_seq">Riesgo · Sequías Severas (S/)</option>
          <option value="mar_comb" selected>Riesgo · combinado (S/)</option>
        </optgroup>
      </select>
    </div>
    <div><label>Agregar tablas por nivel:</label>
      <select id="nivel">
        <option value="dep" selected>Departamento</option>
        <option value="prov">Provincia</option>
      </select>
    </div>
    <div><label>Filtrar provincia (opcional):</label>
      <select id="provfilter"><option value="">— Todas —</option></select>
    </div>
  </div>

  <div class="grid">
    <div id="map" style="height:560px;"></div>
    <div>
      <div class="card" style="margin-top:0;">
        <h3 id="rankTitle">Ranking por departamento</h3>
        <div id="bar" style="height:300px;"></div>
      </div>
      <div class="card">
        <h3>Exposición por entidad</h3>
        <div class="scroll" style="max-height:200px;"><table id="empTbl"></table></div>
      </div>
    </div>
  </div>

  <div class="card">
    <h3 id="tblTitle">Detalle por ubicación</h3>
    <p class="note">Clic en los encabezados para ordenar. La "plata en riesgo" combina el monto
    invertido asignado a cada zona con el score de peligro del CENEPRED (1=menor, 5=mayor).</p>
    <div class="scroll"><table id="dataTbl"></table></div>
  </div>

  <div class="card">
    <h3>Equivalencias de unidades de producción</h3>
    <table>
      <tr><th>Abrev.</th><th>Significado</th><th>Inglés</th><th>Uso típico</th><th>Conversión a TMF</th></tr>
      <tr><td>TMF</td><td>Tonelada Métrica Fina</td><td>Metric Fine Ton</td><td>Cobre, zinc, plomo, molibdeno, hierro, estaño</td><td>× 1</td></tr>
      <tr><td>Grs. f.</td><td>Gramos finos</td><td>Fine grams</td><td>Oro</td><td>÷ 1 000 000</td></tr>
      <tr><td>Kg. f.</td><td>Kilogramos finos</td><td>Fine kilograms</td><td>Plata</td><td>÷ 1 000</td></tr>
    </table>
    <details>
      <summary>Metodología y supuestos</summary>
      <p class="note">
      La exposición de cartera está a nivel de <b>empresa</b> (suma de S/ por instrumento en
      "Valorización de Instrumentos", columna VALOR_VECTOR_LOCAL). Para ubicarla geográficamente
      se reparte entre los distritos donde cada empresa produce, en proporción a su producción en
      <b>TMF</b> (método "<span id="met"></span>"). <b>Cuidado:</b> al unificar todo a TMF, el oro
      (gramos) y la plata (kilogramos) quedan numéricamente pequeños frente a metales base, por lo
      que la asignación tiende a concentrarse en distritos de metales base. El desglose por mineral
      en unidad original (en el Excel y en el tooltip del mapa) permite contrastar esa composición.
      Cambia el método en <code>METODO_ASIGNACION</code> del script si prefieres reparto por
      unidades o equitativo.</p>
      <p class="note"><b>Sector financiero (bancos).</b> Sigue la misma lógica: la exposición de
      cartera a cada grupo bancario (Credicorp, Intercorp Perú, BBVA) sale de "Valorización de
      Instrumentos" y se reparte entre distritos en proporción al <b>volumen de créditos colocados</b>
      por el banco en cada distrito (archivo B-2358-resumen), que cumple el rol que la producción
      tiene en minería. Usa el selector <b>Sector</b> para ver minería, banca o ambos combinados.
      La consolidación de razones sociales por grupo (p. ej. qué líneas suman a "Credicorp") es
      editable en <code>EQUIV_EXPOSICION_FINANCIERO</code> y el script imprime una auditoría de lo
      que sumó a cada grupo.</p>
    </details>
  </div>
</div>

<script>
const PAYLOAD = __PAYLOAD__;
const GEO = __GEOJSON__;
const AFP = PAYLOAD.afp;
const D = PAYLOAD.districts;
document.getElementById('sub').textContent =
  'AFP Integra · ESG · Riesgos Físicos · Generado ' + PAYLOAD.generado +
  ' · Método de asignación: ' + PAYLOAD.metodo;
document.getElementById('met').textContent = PAYLOAD.metodo;

const EMP_SECTOR = PAYLOAD.emp_sector || {};
const SECTORES = PAYLOAD.sectores || [];

const byU = {}; D.forEach(d=>byU[d.ubigeo]=d);
const fmtS = v => 'S/ ' + (v||0).toLocaleString('es-PE',{maximumFractionDigits:0});
const fmtN = v => (v||0).toLocaleString('es-PE',{maximumFractionDigits:2});

const sectorSel = document.getElementById('sector');
function curSector(){ return sectorSel.value; }

// Exposición efectiva del distrito según el sector seleccionado.
function effExp(d){
  const s = curSector();
  if(s==='todos') return d.exposicion||0;
  return (d.exp_sec && d.exp_sec[s]) || 0;
}
function effHasExp(d){ return effExp(d) > 0; }
// Empresas del distrito filtradas por sector.
function effEmpresas(d){
  const s = curSector();
  const all = d.empresas||{};
  if(s==='todos') return all;
  const o={}; Object.entries(all).forEach(([k,v])=>{ if(EMP_SECTOR[k]===s) o[k]=v; });
  return o;
}
// Desglose (minerales / créditos) filtrado por sector.
function effMinerales(d){
  const s = curSector();
  const all = d.minerales||[];
  return s==='todos'? all : all.filter(m=>m.sector===s);
}
// Valor de la métrica para un distrito, respetando el sector.
function metricVal(d, metric){
  if(metric.indexOf('score_')===0) return d[metric];
  const e = effExp(d);
  if(metric==='exposicion') return e;
  const sk = {mar_mm:'score_mm',mar_inu:'score_inu',mar_seq:'score_seq',mar_comb:'score_comb'}[metric];
  const sc = d[sk];
  return sc? e*sc : 0;
}

const METRIC_META = {
  score_mm:{t:'Score Mov. de Masa',score:true},
  score_inu:{t:'Score Inundación',score:true},
  score_seq:{t:'Score Sequías Severas',score:true},
  score_comb:{t:'Score combinado',score:true},
  exposicion:{t:'Exposición asignada (S/)',score:false},
  mar_mm:{t:'Plata en riesgo · Mov. de Masa (S/)',score:false},
  mar_inu:{t:'Plata en riesgo · Inundación (S/)',score:false},
  mar_seq:{t:'Plata en riesgo · Sequías (S/)',score:false},
  mar_comb:{t:'Plata en riesgo · combinado (S/)',score:false},
};

// ---- provincias para el filtro
const provSet = [...new Set(D.map(d=>d.dep+' / '+d.prov))].sort();
const pf = document.getElementById('provfilter');
provSet.forEach(p=>{const o=document.createElement('option');o.value=p;o.textContent=p;pf.appendChild(o);});

function activeDistricts(){
  const f = pf.value;
  return f? D.filter(d=>(d.dep+' / '+d.prov)===f) : D;
}

// ---- KPIs
function renderKPIs(){
  const s=curSector();
  const empSel = Object.entries(PAYLOAD.exposicion_empresa)
    .filter(([k,v])=> s==='todos' || EMP_SECTOR[k]===s);
  const totExp = empSel.reduce((a,b)=>a+b[1],0);
  const nEmp = empSel.filter(([k,v])=>v>0).length;
  const nDist = D.filter(d=>effHasExp(d)).length;
  const nScore = D.filter(d=>d.has_score).length;
  const byDep={}; D.forEach(d=>{const m=metricVal(d,'mar_comb'); if(m>0)byDep[d.dep]=(byDep[d.dep]||0)+m;});
  let topDep='—',topV=-1; Object.entries(byDep).forEach(([k,v])=>{if(v>topV){topV=v;topDep=k;}});
  const scAgg={}; D.forEach(d=>{if(d.score_comb!=null){const a=scAgg[d.dep]||(scAgg[d.dep]={s:0,n:0});a.s+=d.score_comb;a.n++;}});
  let topRiskDep='—',topRiskV=-1;
  Object.entries(scAgg).forEach(([k,o])=>{const m=o.s/o.n; if(m>topRiskV){topRiskV=m;topRiskDep=k;}});
  const etqExp = s==='todos'? 'Exposición total (cartera)' : 'Exposición '+s;
  const k = [
    ['Mayor riesgo físico CENEPRED (depto.)', topRiskDep, 'score prom. '+topRiskV.toFixed(2)+' · panorama nacional'],
    ['Mayor plata en riesgo (depto.)', topDep, fmtS(topV)+' (exposición × score)'],
    [etqExp, fmtS(totExp), nEmp+' entidades · '+nDist+' distritos con plata'],
    ['Cobertura CENEPRED', nScore+' distritos', 'con score de peligro 1–5'],
  ];
  document.getElementById('kpis').innerHTML = k.map(x=>
    `<div class="kpi"><div class="v">${x[1]}</div><div class="l">${x[0]}<br><span style="color:#9aa1b3">${x[2]}</span></div></div>`).join('');
}

// ---- Mapa
function drawMap(){
  const metric = document.getElementById('metric').value;
  const meta = METRIC_META[metric];
  const act = activeDistricts();
  const actSet = new Set(act.map(d=>d.ubigeo));
  const baseLoc=[], baseTxt=[];
  const dataLoc=[], dataZ=[], dataTxt=[];
  GEO.features.forEach(f=>{
    const u=f.properties.IDDIST; const d=byU[u];
    baseLoc.push(u);
    let val=null;
    if(d && actSet.has(u)){
      val = metricVal(d, metric);
      if(!meta.score && !effHasExp(d)) val=null;
      if(val===undefined) val=null;
    }
    let txt;
    if(d && (d.has_score || effHasExp(d))){
      const emp=Object.entries(effEmpresas(d)).sort((a,b)=>b[1]-a[1])
        .map(e=>e[0]+' ('+(EMP_SECTOR[e[0]]||'')+'): '+fmtS(e[1])).join('<br>');
      const mins=effMinerales(d);
      const minMin=mins.filter(m=>m.sector!=='Financiero')
        .map(m=>'· '+m.mineral+': '+fmtN(m.cantidad_original)+' '+m.unidad).join('<br>');
      const minFin=mins.filter(m=>m.sector==='Financiero')
        .map(m=>'· Créditos: '+fmtN(m.cantidad_original)+' '+m.unidad).join('<br>');
      let t=`<b>${d.dist}</b><br>${d.prov}, ${d.dep}<br>`+
        `<b>Score CENEPRED</b> — M.Masa: ${d.score_mm??'s/d'} · Inund.: ${d.score_inu??'s/d'} · Sequías: ${d.score_seq??'s/d'}`;
      if(effHasExp(d)){
        t+=`<br>Exposición asignada: ${fmtS(effExp(d))}`;
        t+=`<br>Plata en riesgo (comb.): ${fmtS(metricVal(d,'mar_comb'))}`;
        if(emp) t+=`<br><b>Entidades:</b><br>${emp}`;
        if(minMin) t+=`<br><b>Producción (unidad original):</b><br>${minMin}`;
        if(minFin) t+=`<br><b>Colocaciones:</b><br>${minFin}`;
      } else {
        t+=`<br><i>sin exposición de cartera</i>`;
      }
      txt=t;
    } else {
      txt=`<b>${f.properties.NOMBDIST}</b><br>${f.properties.NOMBPROV}, ${f.properties.NOMBDEP}<br><i>sin score ni exposición</i>`;
    }
    baseTxt.push(txt);
    if(val!==null){ dataLoc.push(u); dataZ.push(val); dataTxt.push(txt); }
  });
  const colorscale = meta.score
    ? [[0,'#2c7bb6'],[0.25,'#abd9e9'],[0.5,'#ffffbf'],[0.75,'#fdae61'],[1,'#d7191c']]
    : [[0,'#eaf6f8'],[0.5,AFP.cyan],[1,AFP.azul]];
  // Capa base: todo el país en gris claro, con división distrital visible.
  const baseTrace={
    type:'choropleth', geojson:GEO, locations:baseLoc, z:baseLoc.map(()=>0),
    text:baseTxt, hoverinfo:'text', featureidkey:'properties.IDDIST',
    colorscale:[[0,'#eef1f6'],[1,'#eef1f6']], showscale:false,
    marker:{line:{color:'#b9c0cf',width:0.4}},
  };
  // Capa de datos: solo distritos con valor, coloreados y resaltados con borde blanco.
  const dataTrace={
    type:'choropleth', geojson:GEO, locations:dataLoc, z:dataZ,
    text:dataTxt, hoverinfo:'text', featureidkey:'properties.IDDIST', colorscale,
    zmin: meta.score?1:undefined, zmax: meta.score?5:undefined,
    marker:{line:{color:'#ffffff',width:0.6}},
    colorbar:{title:{text:meta.score?'Score':'S/',side:'right'},thickness:12,len:.8},
  };
  const data=[baseTrace, dataTrace];
  const subt = meta.score
    ? 'Panorama nacional — '+act.filter(d=>d.has_score).length+' distritos con score'
    : 'Distritos con exposición'+(curSector()!=='todos'?' ('+curSector()+')':'')+
      ' ('+act.filter(d=>effHasExp(d)).length+')';
  const layout={
    geo:{fitbounds:'locations',visible:false,bgcolor:'rgba(0,0,0,0)'},
    margin:{l:0,r:0,t:28,b:0}, paper_bgcolor:'#fff',
    title:{text:meta.t+'  <span style="font-size:11px;color:#6b7280">— '+subt+'</span>',
           font:{size:14,color:AFP.azul},x:0.02,y:0.985},
  };
  Plotly.react('map',data,layout,{displayModeBar:false,responsive:true});
}

// ---- Ranking (bar) por nivel
function drawBar(){
  const metric=document.getElementById('metric').value;
  const meta=METRIC_META[metric];
  const nivel=document.getElementById('nivel').value;
  const act=activeDistricts();
  const agg={};
  act.forEach(d=>{
    const key = nivel==='dep'? d.dep : (d.dep+' / '+d.prov);
    if(!agg[key]) agg[key]={sum:0,ssc:0,ncnt:0};
    if(meta.score){
      const v=d[metric]; if(v!=null){agg[key].ssc+=v; agg[key].ncnt+=1;}
    } else {
      if(effHasExp(d)) agg[key].sum += metricVal(d, metric)||0;
    }
  });
  let rows=Object.entries(agg).map(([k,o])=>[k, meta.score? (o.ncnt? o.ssc/o.ncnt:0):o.sum]);
  rows=rows.filter(r=>r[1]>0).sort((a,b)=>b[1]-a[1]).slice(0,12).reverse();
  document.getElementById('rankTitle').textContent =
    (nivel==='dep'?'Ranking por departamento — ':'Ranking por provincia — ')+meta.t;
  const data=[{type:'bar',orientation:'h',x:rows.map(r=>r[1]),y:rows.map(r=>r[0]),
    marker:{color: meta.score? '#d7191c' : AFP.azul},
    text:rows.map(r=>meta.score? r[1].toFixed(2): fmtS(r[1])),textposition:'auto',
    hovertemplate:'%{y}<br>'+(meta.score?'score prom. %{x:.2f}':'S/ %{x:,.0f}')+'<extra></extra>'}];
  const layout={margin:{l:140,r:20,t:6,b:24},paper_bgcolor:'#fff',plot_bgcolor:'#fff',
    xaxis:{showgrid:false,zeroline:false,range:meta.score?[0,5]:undefined},yaxis:{automargin:true}};
  Plotly.react('bar',data,layout,{displayModeBar:false,responsive:true});
}

// ---- Tablas
let sortState={col:null,dir:1};
function drawTable(){
  const metric=document.getElementById('metric').value;
  const act=activeDistricts();
  const cols=[['dep','Departamento',0],['prov','Provincia',0],['dist','Distrito',0],
    ['exposicion','Exposición (S/)',1],['score_mm','M.Masa',1],['score_inu','Inund.',1],
    ['score_seq','Sequías',1],['mar_comb','Plata en riesgo comb. (S/)',1]];
  let rows=act.slice();
  const sc = sortState.col || 'mar_comb';
  const valOf=(d,c)=>(c==='exposicion'||c==='mar_comb')? metricVal(d,c) : (d[c]??-1);
  rows.sort((a,b)=>{const x=valOf(a,sc),y=valOf(b,sc); return (x<y?-1:x>y?1:0)*sortState.dir;});
  if(!sortState.col) rows.reverse(); // default desc by mar_comb
  let h='<tr>'+cols.map(c=>`<th class="${c[2]?'num':''}" data-c="${c[0]}">${c[1]}</th>`).join('')+'</tr>';
  let body=rows.slice(0,400).map(d=>'<tr>'+cols.map(c=>{
    let v=d[c[0]];
    if(c[0]==='exposicion'||c[0]==='mar_comb'){ v=fmtS(metricVal(d,c[0])); }
    else if(c[2]){ v = (v??'s/d'); }
    return `<td class="${c[2]?'num':''}">${v}</td>`;
  }).join('')+'</tr>').join('');
  const t=document.getElementById('dataTbl'); t.innerHTML=h+body;
  t.querySelectorAll('th').forEach(th=>th.onclick=()=>{
    const c=th.dataset.c; sortState.dir=(sortState.col===c)?-sortState.dir:-1; sortState.col=c; drawTable();
  });
  document.getElementById('tblTitle').textContent =
    'Detalle por distrito'+(pf.value?(' — '+pf.value):'')+' ('+rows.length+' distritos)';
}

function drawEmpTable(){
  const s=curSector();
  let e=Object.entries(PAYLOAD.exposicion_empresa)
    .filter(([k,v])=> s==='todos' || EMP_SECTOR[k]===s)
    .sort((a,b)=>b[1]-a[1]);
  const tot=e.reduce((a,b)=>a+b[1],0)||1;
  let h='<tr><th>Entidad</th><th>Sector</th><th class="num">Exposición (S/)</th><th class="num">%</th></tr>';
  h+=e.map(([k,v])=>`<tr><td>${k}</td><td>${EMP_SECTOR[k]||''}</td><td class="num">${fmtS(v)}</td><td class="num">${(100*v/tot).toFixed(1)}%</td></tr>`).join('');
  document.getElementById('empTbl').innerHTML=h;
}

function redrawAll(){renderKPIs();drawEmpTable();drawMap();drawBar();drawTable();}
['sector','metric','nivel','provfilter'].forEach(id=>document.getElementById(id).addEventListener('change',redrawAll));
renderKPIs();drawEmpTable();drawMap();drawBar();drawTable();
</script>
</body></html>
"""


# ==============================================================================
# 10) MAIN
# ==============================================================================

def main():
    log("=" * 60)
    log("ANÁLISIS DE RIESGOS FÍSICOS DE CARTERA (MINERÍA + FINANCIERO) — INICIO")
    log("=" * 60)

    prod_path = find_file(PROD_DIR, PROD_PATTERNS, "PRODUCCIÓN MINERA")
    risk_path = find_file(RISK_DIR, RISK_PATTERNS, "RIESGOS FÍSICOS (CENEPRED + Valorización)")
    log(f"Producción : {prod_path}")
    log(f"Riesgos    : {risk_path}")

    geojson = load_geojson()
    geo_idx = build_geo_index(geojson)

    sheet_cen = pick_sheet(risk_path, SHEET_CENEPRED_HINTS, "CENEPRED")
    sheet_val = pick_sheet(risk_path, SHEET_VALORIZ_HINTS, "Valorización de Instrumentos")
    scores = load_cenepred(risk_path, sheet_cen)

    # ---------- SECTOR MINERÍA ----------
    df_prod, _ = load_produccion(prod_path)
    exp = load_exposicion(risk_path, sheet_val)
    districts_min, df_assign, um, detalle, no_map, dfm = build_model(
        df_prod, exp, scores, geo_idx, EMPRESAS_CANONICAS, SECTOR_MINERIA)

    # mapa empresa -> sector y exposición combinada (arranca con minería)
    emp_sector = {e: SECTOR_MINERIA for e in EMPRESAS_CANONICAS}
    exp_comb = dict(exp)
    listas = [districts_min]

    # acumuladores para el Excel (mantienen separación por nombre de empresa)
    df_assign_all = df_assign.copy()
    um_all = dict(um)
    detalle_all = dict(detalle)
    no_map_all = no_map.copy()
    universo_all = list(EMPRESAS_CANONICAS)

    # ---------- SECTOR FINANCIERO (BANCOS) ----------
    if INCLUIR_FINANCIERO:
        try:
            fin_path = find_file(FIN_DIR, FIN_PATTERNS, "RESUMEN DE BANCOS (créditos)")
            log(f"Financiero : {fin_path}")
            df_fin, _ = load_financiero(fin_path)
            exp_fin = load_exposicion_financiero(risk_path, sheet_val)
            districts_fin, df_assign_f, um_f, detalle_f, no_map_f, _ = build_model(
                df_fin, exp_fin, scores, geo_idx, EMPRESAS_FIN_CANONICAS, SECTOR_FINANCIERO)

            listas.append(districts_fin)
            for e in EMPRESAS_FIN_CANONICAS:
                emp_sector[e] = SECTOR_FINANCIERO
            exp_comb.update(exp_fin)
            df_assign_all = pd.concat([df_assign_all, df_assign_f], ignore_index=True)
            um_all.update(um_f)
            detalle_all.update(detalle_f)
            no_map_all = pd.concat([no_map_all, no_map_f], ignore_index=True)
            universo_all += list(EMPRESAS_FIN_CANONICAS)
        except SystemExit:
            raise
        except Exception as e:
            log(f"No se pudo procesar el sector financiero: {e}\n"
                f"      -> El dashboard se generará solo con minería. Revisa FIN_DIR / "
                f"FIN_PATTERNS / nombres en CONFIG.", "WARN")

    # ---------- FUSIÓN POR DISTRITO ----------
    districts = merge_sectores(listas)

    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    xlsx_out = os.path.join(OUT_DIR, f"Riesgos_Fisicos_Cartera_{stamp}.xlsx")
    html_out = os.path.join(OUT_DIR, f"Mapa_Riesgos_Fisicos_{stamp}.html")

    export_excel(xlsx_out, df_assign_all, exp_comb, scores, geo_idx, um_all,
                 detalle_all, districts, no_map_all,
                 universo=universo_all, emp_sector=emp_sector)
    log(f"Excel generado : {xlsx_out}")
    export_html(html_out, geojson, districts, exp_comb, scores, geo_idx,
                emp_sector=emp_sector)
    log(f"HTML generado  : {html_out}")

    log("=" * 60)
    log("PROCESO COMPLETADO")
    log("=" * 60)
    return xlsx_out, html_out

    log("=" * 60)
    log("PROCESO COMPLETADO")
    log("=" * 60)
    return xlsx_out, html_out


if __name__ == "__main__":
    main()