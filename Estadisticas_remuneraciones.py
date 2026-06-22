# -*- coding: utf-8 -*-
"""
Estadisticas_remuneraciones.py
================================================================================
Estimacion bottom-up de la remuneracion del directorio de las empresas del
portafolio, a partir de dos insumos:

  FUENTE 1  _consolidado_remuneracion_directorio.xlsx
            (lo genera el script Extrae_Estadisticas_de_Directores_SMV.py:
             una hoja por empresa, con bloques por ejercicio que contienen el
             resumen del directorio, los comites y la matriz de participacion).

  FUENTE 2  Remuneraciones_Directorio_{YYYYMMDD}_{HHMM}.xlsx
            (archivo de dietas; se toma siempre el mas reciente por timestamp).

Genera dos salidas:

  OUTPUT 1  Una hoja nueva en Remuneracion.xlsx (sin tocar nada existente),
            cuyo nombre es la fecha de ejecucion en formato DD-MM-YYYY.

  OUTPUT 2  Un dashboard HTML interactivo autocontenido.

IMPORTANTE
----------
La remuneracion calculada es una ESTIMACION APROXIMADA. Los reportes SMV no
informan el monto pagado por sesion a cada miembro ni la asistencia individual
a cada comite, por lo que el calculo multiplica las sesiones (de directorio y
de comites) por la tarifa por sesion vigente extraida del archivo de dietas.
Ver metodologia detallada en Extrae_Estadisticas_de_Directores_SMV.py.

Dependencias: pandas, openpyxl  (+ glob, pathlib, datetime de la libreria estandar)
Sin xlwings ni win32com.
"""

# %% ===========================================================================
#    PARAMETROS  (editar aqui las rutas y constantes)
# ==============================================================================

import os
import re
import glob
import json
import shutil
import unicodedata
from pathlib import Path
from datetime import datetime
from collections import defaultdict, OrderedDict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- Rutas fijas (entorno de produccion) -------------------------------------
RUTA_BASE = Path(
    r"C:\Users\usuario\OneDrive\Desktop\AFP INTEGRA\Docs privados\Otros"
    r"\Remuneracion Directorio - Bottom Up"
)
RUTA_CONSOLIDADO = RUTA_BASE / "Remuneraciones" / "_consolidado_remuneracion_directorio.xlsx"
RUTA_MAESTRO     = RUTA_BASE / "Remuneracion.xlsx"
# El archivo de dietas tiene nombre variable; se busca por patron en RUTA_BASE.
PATRON_DIETAS    = "Remuneraciones_Directorio_*.xlsx"

# --- Fallback de ejecucion ----------------------------------------------------
# Si la ruta de produccion no existe (p.ej. al probar el script en otra maquina),
# se usa la carpeta donde reside este .py. Asi el script es portatil sin perder
# las rutas oficiales de arriba.
if not RUTA_BASE.exists():
    _AQUI = Path(__file__).resolve().parent
    RUTA_BASE = _AQUI
    # Se intenta primero la subcarpeta "Remuneraciones"; si no, la carpeta actual.
    _cons = _AQUI / "Remuneraciones" / "_consolidado_remuneracion_directorio.xlsx"
    RUTA_CONSOLIDADO = _cons if _cons.exists() else _AQUI / "_consolidado_remuneracion_directorio.xlsx"
    RUTA_MAESTRO = _AQUI / "Remuneracion.xlsx"

# --- Sectores y orden de empresas (CONSTANTE EDITABLE) ------------------------
# Si en el futuro hay mas empresas o cambia el orden, basta editar esta lista:
# la hoja Excel y el dashboard se adaptan solos.
SECTORES = OrderedDict([
    ("Mineras",      ["Buenaventura", "Cerro Verde", "Minsur", "Volcan", "Nexa Perú", "Hochshild"]),
    ("Utilities",    ["Orygen", "Pluz Energia", "Engie"]),
    ("Construcción", ["Ferreycorp", "Aenza", "UNACEM", "Pacasmayo"]),
    ("Consumo",      ["InRetail", "Alicorp"]),
    ("Financieras",  ["Credicorp", "IFS", "BBVA"]),
])

# --- Universo canonico del portafolio ----------------------------------------
PORTAFOLIO_CANONICO = [
    "Cerro Verde", "Pacasmayo", "Aenza", "Ferreycorp", "Engie", "Volcan", "BBVA",
    "Pluz Energia", "Puerto Chancay", "Orygen", "Buenaventura", "Alicorp", "UNACEM",
    "Credicorp", "Nexa Perú", "InRetail", "Auna", "IFS", "Minsur", "Hudbay",
    "Hunt Oil", "Hermes", "Colegios Peruanos", "Casa Andina", "Inca Rail", "Intursa",
    "Jockey Plaza", "Lima Expresa", "Orazen", "Primax", "Rutas de Lima", "Tecsup",
    "Hochshild",
]

# --- Supuestos del modelo -----------------------------------------------------
# Si aplica_modelo_por_sesion == "SÍ" se asume que la empresa paga solo por las
# sesiones a las que el director ASISTIO; en caso contrario se usan las
# sesiones CONVOCADAS.
PAGA_SOLO_ASISTENCIA = True
# Ante varios montos "Por sesion" del mismo periodo (distintos niveles de
# director), criterio de desempate para la tarifa de referencia: "max" | "min".
CRITERIO_TARIFA_EMPATE = "max"
MONEDA_POR_DEFECTO = "PEN"   # cuando el archivo de dietas no informa la moneda

# --- Paleta AFP Integra -------------------------------------------------------
AZUL    = "1F3864"   # encabezados principales
CELESTE = "4472C4"   # encabezados de sector
FONDO   = "D9E1F2"   # fondo suave (fila de anios)
BORDE   = "BFBFBF"   # bordes
GRIS    = "808080"   # nota al pie
FUENTE_NOMBRE = "Aptos Narrow"   # misma fuente que el Consolidado del maestro

# %% ===========================================================================
#    UTILIDADES
# ==============================================================================

def normalizar(texto):
    """minusculas, sin tildes, sin texto entre corchetes ni sufijos legales."""
    if texto is None:
        return ""
    s = str(texto)
    s = re.sub(r"\[.*?\]", " ", s)                       # quita "[SIN MATCH ...]"
    s = "".join(c for c in unicodedata.normalize("NFD", s)
                if unicodedata.category(c) != "Mn")       # quita tildes
    s = s.lower()
    s = re.sub(r"[.,]", " ", s)
    # sufijos legales al final
    s = re.sub(r"\b(s\s*a\s*a|s\s*a|inc|ltd|sab|saa)\b", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

# Diccionario normalizado -> nombre canonico
CANON_POR_NORM = {normalizar(c): c for c in PORTAFOLIO_CANONICO}

def a_canonico(nombre_empresa):
    """Devuelve el nombre canonico del portafolio o None si esta fuera de el."""
    return CANON_POR_NORM.get(normalizar(nombre_empresa))

_TRIMESTRE = {"1q": 1, "2q": 2, "3q": 3, "4q": 4}
def orden_trimestre(t):
    if t is None:
        return 0
    return _TRIMESTRE.get(str(t).strip().lower(), 0)

def es_por_sesion(frecuencia):
    return "por sesion" in normalizar(frecuencia)

def es_numero(x):
    return isinstance(x, (int, float)) and not isinstance(x, bool)

_SES_RE = re.compile(r"\((\d+)\s*ses", re.IGNORECASE)
def sesiones_en_encabezado(texto):
    """Extrae el N de '... (N ses)' del encabezado de columna de comite."""
    m = _SES_RE.search(str(texto))
    return int(m.group(1)) if m else None


# %% ===========================================================================
#    FUENTE 1  ·  Lectura del consolidado
# ==============================================================================

_HDR_BLOQUE = re.compile(r"^(.*?)\s+·\s+EJERCICIO\s+(\S+)\s+·")

def leer_consolidado(ruta):
    """
    Devuelve:
      registros : lista de dicts (uno por empresa+ejercicio valido) con
                  empresa_canonica, ejercicio (int), n_directores, n_sesiones_dir,
                  aplica_modelo ("SÍ"/"NO"), comites [(nombre,n_ses)],
                  directores [dict(nombre, ses_asistio, ses_convoco,
                                   comites_participa [(nombre_col,n_ses)])]
      fuera_portafolio : set de nombres de empresa fuera del portafolio
      hubo_ejercicio_none : bool
    """
    wb = openpyxl.load_workbook(ruta, data_only=True)
    omitir = {"Índice", "Indice", "Incidencias"}

    registros = []
    fuera_portafolio = set()
    hubo_ejercicio_none = False

    for hoja in wb.sheetnames:
        if hoja in omitir:
            continue
        ws = wb[hoja]

        # 1) localizar todos los bloques "Empresa · EJERCICIO yyyy · ..."
        bloques = []
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if isinstance(v, str):
                m = _HDR_BLOQUE.match(v)
                if m:
                    bloques.append((r, m.group(1).strip(), m.group(2).strip()))

        for bi, (fila_ini, empresa_raw, ejercicio_raw) in enumerate(bloques):
            fila_fin = bloques[bi + 1][0] - 1 if bi + 1 < len(bloques) else ws.max_row

            # registro sin anio detectado -> se ignora con advertencia
            if ejercicio_raw == "None" or ejercicio_raw is None:
                hubo_ejercicio_none = True
                print(f"  [AVISO] Registro con ejercicio=None ignorado: "
                      f"'{empresa_raw}' (hoja {hoja}).")
                continue

            canonico = a_canonico(empresa_raw)
            if canonico is None:
                fuera_portafolio.add(empresa_raw)
                continue  # fuera del portafolio: se excluye de los outputs

            try:
                ejercicio = int(ejercicio_raw)
            except (TypeError, ValueError):
                hubo_ejercicio_none = True
                print(f"  [AVISO] Ejercicio no numerico '{ejercicio_raw}' "
                      f"ignorado ({empresa_raw}).")
                continue

            n_dir = n_ses = aplica = None
            comites = []
            directores = []

            r = fila_ini
            while r <= fila_fin:
                a = ws.cell(r, 1).value
                if isinstance(a, str) and a.startswith("1. Resumen"):
                    # encabezado en r+1, datos en r+2
                    fila_datos = r + 2
                    n_dir  = ws.cell(fila_datos, 1).value
                    n_ses  = ws.cell(fila_datos, 4).value
                    aplica = ws.cell(fila_datos, 8).value  # columna 8 = "¿Aplica modelo por sesion?"

                elif isinstance(a, str) and a.startswith("3. Comit"):
                    rr = r + 2  # salta encabezado de la tabla de comites
                    while rr <= fila_fin:
                        nombre_com = ws.cell(rr, 1).value
                        if nombre_com is None or (isinstance(nombre_com, str)
                                                  and nombre_com.startswith("4.")):
                            break
                        n_ses_com = ws.cell(rr, 3).value
                        try:
                            n_ses_com = int(n_ses_com)
                        except (TypeError, ValueError):
                            n_ses_com = 0
                        comites.append((str(nombre_com).strip(), n_ses_com))
                        rr += 1

                elif isinstance(a, str) and a.startswith("4. Matriz"):
                    fila_hdr = r + 1
                    col_asistio = col_convoco = None
                    cols_comite = []   # (col, nombre_col, n_ses)
                    for c in range(1, ws.max_column + 1):
                        h = ws.cell(fila_hdr, c).value
                        if not isinstance(h, str):
                            continue
                        hl = normalizar(h)
                        if "asistio" in hl:
                            col_asistio = c
                        elif "convoc" in hl:
                            col_convoco = c
                        else:
                            n = sesiones_en_encabezado(h)
                            if n is not None:           # solo columnas de comite
                                cols_comite.append((c, h.strip(), n))
                        # la columna "Σ ses. comites" no tiene "(N ses)" -> se ignora

                    rr = fila_hdr + 1
                    while rr <= fila_fin:
                        nombre = ws.cell(rr, 1).value
                        if nombre is None or (isinstance(nombre, str)
                                              and (nombre.startswith("Nota")
                                                   or nombre.startswith("4."))):
                            break
                        asistio = ws.cell(rr, col_asistio).value if col_asistio else None
                        convoco = ws.cell(rr, col_convoco).value if col_convoco else None
                        participa = []
                        for (c, nombre_col, n_ses_col) in cols_comite:
                            celda = ws.cell(rr, c).value
                            if celda is not None:       # participa en ese comite
                                participa.append((nombre_col, n_ses_col))
                        directores.append(dict(
                            nombre=str(nombre).strip(),
                            ses_asistio=asistio if es_numero(asistio) else 0,
                            ses_convoco=convoco if es_numero(convoco) else 0,
                            comites_participa=participa,
                        ))
                        rr += 1
                    r = rr
                    continue
                r += 1

            registros.append(dict(
                empresa=canonico,
                hoja=hoja,
                ejercicio=ejercicio,
                n_directores=int(n_dir) if es_numero(n_dir) else None,
                n_sesiones_dir=int(n_ses) if es_numero(n_ses) else 0,
                aplica_modelo=(str(aplica).strip().upper() if aplica is not None else "NO"),
                comites=comites,
                directores=directores,
            ))

    return registros, fuera_portafolio, hubo_ejercicio_none


# %% ===========================================================================
#    FUENTE 2  ·  Lectura del archivo de dietas (tarifas por sesion)
# ==============================================================================

def hallar_archivo_dietas(carpeta, patron):
    """Devuelve la ruta del archivo de dietas mas reciente (mayor timestamp)."""
    candidatos = list(Path(carpeta).glob(patron))
    if not candidatos:
        return None

    def clave(p):
        # extrae YYYYMMDD y HHMM del nombre; si no hay, usa la fecha de modificacion
        m = re.search(r"(\d{8})_(\d{4})", p.name)
        if m:
            return (m.group(1), m.group(2))
        return (datetime.fromtimestamp(p.stat().st_mtime).strftime("%Y%m%d"), "0000")

    return max(candidatos, key=clave)


def _filas_dietas(wb):
    """Aplana todas las filas del archivo de dietas en dicts homogeneos."""
    filas = []
    for hoja in wb.sheetnames:
        if normalizar(hoja) == "diagnostico":
            continue
        ws = wb[hoja]
        encabezado = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        idx = {h: i for i, h in enumerate(encabezado) if h}

        def buscar(*claves):
            for k in claves:
                if k in idx:
                    return idx[k]
            return None

        i_emp  = buscar("Empresa")
        i_tipo = buscar("Tipo")
        i_con  = buscar("Concepto")
        i_mon  = buscar("Monto")
        i_cur  = buscar("Moneda")
        i_frec = buscar("Frecuencia")
        i_anio = buscar("Anio", "Año")
        i_tri  = buscar("Trimestre")

        for r in range(2, ws.max_row + 1):
            vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            def g(i):
                return vals[i] if i is not None and i < len(vals) else None
            empresa = g(i_emp) if i_emp is not None else hoja
            filas.append(dict(
                empresa=empresa,
                es_resumen=(normalizar(hoja) == "resumen"),
                tipo=g(i_tipo),
                concepto=g(i_con),
                monto=g(i_mon),
                moneda=g(i_cur),
                frecuencia=g(i_frec),
                anio=g(i_anio),
                trimestre=g(i_tri),
            ))
    return filas


def _elegir(candidatos):
    """De una lista de filas 'Por sesion' con monto valido, elige la vigente."""
    if not candidatos:
        return None
    candidatos = sorted(
        candidatos,
        key=lambda r: ((r["anio"] or 0), orden_trimestre(r["trimestre"]),
                       r["monto"] if CRITERIO_TARIFA_EMPATE == "max" else -r["monto"]),
    )
    return candidatos[-1]


def leer_tarifas(ruta):
    """
    Devuelve dict canonico -> dict(tarifa_dir, tarifa_com, moneda)
    y un set de empresas del archivo de dietas fuera del portafolio.
    """
    wb = openpyxl.load_workbook(ruta, data_only=True)
    filas = _filas_dietas(wb)

    por_empresa = defaultdict(list)
    fuera = set()
    for f in filas:
        canon = a_canonico(f["empresa"])
        if canon is None:
            if f["empresa"]:
                fuera.add(str(f["empresa"]))
            continue
        por_empresa[canon].append(f)

    tarifas = {}
    for canon, regs in por_empresa.items():
        por_sesion = [r for r in regs
                      if es_por_sesion(r["frecuencia"]) and es_numero(r["monto"]) and r["monto"] > 0]

        # --- tarifa de directorio ---
        # 1) preferir la hoja RESUMEN (cifra curada por el analista)
        pool = [r for r in por_sesion if r["es_resumen"]]
        # 2) si no hay, filas tipo "Directorio" de hojas de detalle
        if not pool:
            pool = [r for r in por_sesion if normalizar(r["tipo"]) == "directorio"]
        # 3) si no hay, cualquier fila "Por sesion"
        if not pool:
            pool = por_sesion
        fila_dir = _elegir(pool)

        # --- tarifa de comite (monto explicito etiquetado para comites) ---
        pool_com = [r for r in por_sesion if normalizar(r["tipo"]) == "comite"]
        fila_com = _elegir(pool_com)

        if fila_dir is None and fila_com is None:
            continue  # sin tarifa utilizable

        tarifa_dir = fila_dir["monto"] if fila_dir else None
        moneda = (fila_dir["moneda"] if fila_dir and fila_dir["moneda"]
                  else (fila_com["moneda"] if fila_com and fila_com["moneda"] else None))
        moneda = moneda or MONEDA_POR_DEFECTO

        if tarifa_dir is None:                       # solo habia tarifa de comite
            tarifa_dir = fila_com["monto"]
        tarifa_com = fila_com["monto"] if fila_com else tarifa_dir   # supuesto: comite = directorio

        tarifas[canon] = dict(
            tarifa_dir=float(tarifa_dir),
            tarifa_com=float(tarifa_com),
            moneda=str(moneda).strip().upper(),
        )

    return tarifas, fuera


# %% ===========================================================================
#    LOGICA DE CALCULO (bottom-up)
# ==============================================================================

def calcular(registros, tarifas):
    """
    A partir de los registros del consolidado y las tarifas por sesion,
    calcula la remuneracion estimada por empresa y ejercicio.

    Devuelve dict[empresa][ejercicio] = dict(
        remuneracion_total, n_directores, moneda,
        comites_n, comites_directores_participan, comites_sesiones_total
    )
    y el conjunto de anios disponibles.
    """
    resultados = defaultdict(dict)
    anios = set()

    for rec in registros:
        emp = rec["empresa"]
        anio = rec["ejercicio"]
        anios.add(anio)

        tar = tarifas.get(emp)
        tarifa_dir = tar["tarifa_dir"] if tar else None
        tarifa_com = tar["tarifa_com"] if tar else None
        moneda = tar["moneda"] if tar else MONEDA_POR_DEFECTO

        # --- estadisticas de comites (independientes de la tarifa) ---
        comites_n = len(rec["comites"])
        comites_sesiones_total = sum(n for _, n in rec["comites"])
        directores_en_comite = sum(
            1 for d in rec["directores"] if len(d["comites_participa"]) > 0
        )

        remuneracion_total = None
        if tarifa_dir is not None:
            paga_asistencia = (rec["aplica_modelo"] == "SÍ") and PAGA_SOLO_ASISTENCIA
            total = 0.0
            for d in rec["directores"]:
                ses_pagables = d["ses_asistio"] if paga_asistencia else d["ses_convoco"]
                rem_directorio = tarifa_dir * (ses_pagables or 0)
                # Sigma de sesiones de comite donde el director participa
                ses_comites = sum(n_ses for (_, n_ses) in d["comites_participa"])
                rem_comites = tarifa_com * ses_comites
                total += rem_directorio + rem_comites
            remuneracion_total = total

        resultados[emp][anio] = dict(
            remuneracion_total=remuneracion_total,
            n_directores=rec["n_directores"],
            moneda=moneda,
            comites_n=comites_n,
            comites_directores_participan=directores_en_comite,
            comites_sesiones_total=comites_sesiones_total,
        )

    return resultados, sorted(anios)


# %% ===========================================================================
#    OUTPUT 1  ·  Hoja nueva en Remuneracion.xlsx
# ==============================================================================

def _estilos():
    delgado = Side(style="thin", color=BORDE)
    borde = Border(left=delgado, right=delgado, top=delgado, bottom=delgado)
    f_titulo = Font(name=FUENTE_NOMBRE, bold=True, color="FFFFFF", size=11)
    f_sector = Font(name=FUENTE_NOMBRE, bold=True, color="FFFFFF", size=11)
    f_normal = Font(name=FUENTE_NOMBRE, size=11)
    f_nota = Font(name=FUENTE_NOMBRE, italic=True, color=GRIS, size=9)
    fill_azul = PatternFill("solid", fgColor=AZUL)
    fill_celeste = PatternFill("solid", fgColor=CELESTE)
    fill_fondo = PatternFill("solid", fgColor=FONDO)
    return dict(borde=borde, f_titulo=f_titulo, f_sector=f_sector, f_normal=f_normal,
                f_nota=f_nota, fill_azul=fill_azul, fill_celeste=fill_celeste,
                fill_fondo=fill_fondo)


def escribir_hoja_excel(ruta_maestro, resultados, anio_n1, anio_n2):
    """
    Inserta una hoja nueva (nombre = fecha de ejecucion DD-MM-YYYY) replicando
    la estructura E3:J30 del 'Consolidado'. No modifica ninguna hoja existente.
    Devuelve el nombre de la hoja creada.
    """
    # copia de seguridad antes de tocar el maestro
    try:
        respaldo = ruta_maestro.with_name(
            ruta_maestro.stem + "_backup_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".xlsx")
        shutil.copy2(ruta_maestro, respaldo)
        print(f"  Respaldo del maestro: {respaldo.name}")
    except Exception as e:
        print(f"  [AVISO] No se pudo crear respaldo del maestro: {e}")

    wb = openpyxl.load_workbook(ruta_maestro, keep_links=True)

    nombre_hoja = datetime.now().strftime("%d-%m-%Y")
    base = nombre_hoja
    n = 1
    while nombre_hoja in wb.sheetnames:   # evita colision si ya existe la del dia
        n += 1
        nombre_hoja = f"{base} ({n})"

    ws = wb.create_sheet(title=nombre_hoja)   # se inserta al final
    st = _estilos()
    fmt_miles = '_-* #,##0_-;\\-* #,##0_-;_-* "-"??_-;_-@_-'

    C_EMP, C_MON, C_R1, C_R2, C_D1, C_D2 = 5, 6, 7, 8, 9, 10   # E..J
    FILA_G, FILA_A = 3, 4

    # --- Fila 3: encabezados de grupo (mergeados) ---
    ws.cell(FILA_G, C_R1, "Remuneracion Directorio (000)")
    ws.merge_cells(start_row=FILA_G, start_column=C_R1, end_row=FILA_G, end_column=C_R2)
    ws.cell(FILA_G, C_D1, "# Directores")
    ws.merge_cells(start_row=FILA_G, start_column=C_D1, end_row=FILA_G, end_column=C_D2)
    for c in (C_R1, C_D1):
        cell = ws.cell(FILA_G, c)
        cell.font = st["f_titulo"]; cell.fill = st["fill_azul"]
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Fila 4: encabezados de anios ---
    encabezados = [(C_EMP, "Empresa"), (C_MON, "Moneda"),
                   (C_R1, anio_n1), (C_R2, anio_n2),
                   (C_D1, anio_n1), (C_D2, anio_n2)]
    for c, txt in encabezados:
        cell = ws.cell(FILA_A, c, txt)
        cell.font = st["f_titulo"]
        cell.fill = st["fill_azul"] if c in (C_EMP, C_MON) else st["fill_fondo"]
        if c in (C_EMP, C_MON):
            cell.font = st["f_titulo"]
        else:
            cell.font = Font(name=FUENTE_NOMBRE, bold=True, color=AZUL, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = st["borde"]

    # --- Filas de datos por sector ---
    fila = FILA_A + 1
    for sector, empresas in SECTORES.items():
        # fila de sector
        cell = ws.cell(fila, C_EMP, sector)
        cell.font = st["f_sector"]; cell.fill = st["fill_celeste"]
        for c in range(C_EMP, C_D2 + 1):
            cc = ws.cell(fila, c)
            cc.fill = st["fill_celeste"]; cc.border = st["borde"]
        fila += 1

        for emp in empresas:
            datos = resultados.get(emp, {})
            d1 = datos.get(anio_n1); d2 = datos.get(anio_n2)
            moneda = (d1 or d2 or {}).get("moneda", "-") if (d1 or d2) else "-"

            def rem_miles(d):
                if d and d.get("remuneracion_total") is not None:
                    return round(d["remuneracion_total"] / 1000.0)
                return "-"
            def ndir(d):
                if d and d.get("n_directores") is not None:
                    return d["n_directores"]
                return "-"

            ws.cell(fila, C_EMP, emp).font = st["f_normal"]
            ws.cell(fila, C_MON, moneda).font = st["f_normal"]
            ws.cell(fila, C_MON).alignment = Alignment(horizontal="center")
            valores = [(C_R1, rem_miles(d1)), (C_R2, rem_miles(d2)),
                       (C_D1, ndir(d1)), (C_D2, ndir(d2))]
            for c, v in valores:
                cc = ws.cell(fila, c, v)
                cc.font = st["f_normal"]
                if c in (C_R1, C_R2):
                    cc.number_format = fmt_miles
                    cc.alignment = Alignment(horizontal="right")
                else:
                    cc.alignment = Alignment(horizontal="center")
            for c in range(C_EMP, C_D2 + 1):
                ws.cell(fila, c).border = st["borde"]
            fila += 1

    # --- Nota al pie (dos filas debajo de los datos) ---
    fila_nota = fila + 1
    nota = ("Estimación bottom-up. Ver metodología en "
            "Extrae_Estadisticas_de_Directores_SMV.py")
    cell = ws.cell(fila_nota, C_EMP, nota)
    cell.font = st["f_nota"]
    ws.merge_cells(start_row=fila_nota, start_column=C_EMP, end_row=fila_nota, end_column=C_D2)

    # anchos de columna
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 9
    for col in ("G", "H"):
        ws.column_dimensions[col].width = 13
    for col in ("I", "J"):
        ws.column_dimensions[col].width = 9

    wb.save(ruta_maestro)
    return nombre_hoja


# %% ===========================================================================
#    OUTPUT 2  ·  Dashboard HTML interactivo (autocontenido)
# ==============================================================================

def _construir_payload(resultados, anios_disponibles, anio_n1, anio_n2):
    """Arma la estructura de datos que consume el JS del dashboard."""
    sectores = []
    empresas = OrderedDict()
    comites = OrderedDict()
    sector_de = {}
    for sector, emps in SECTORES.items():
        sectores.append({"nombre": sector, "empresas": emps})
        for e in emps:
            sector_de[e] = sector

    for sector, emps in SECTORES.items():
        for emp in emps:
            datos = resultados.get(emp, {})
            por_anio = {}
            for anio in anios_disponibles:
                d = datos.get(anio)
                if not d:
                    por_anio[str(anio)] = {"rem": None, "ndir": None, "prom": None}
                    continue
                rem = (round(d["remuneracion_total"] / 1000.0, 1)
                       if d["remuneracion_total"] is not None else None)
                ndir = d["n_directores"]
                prom = (round(rem / ndir, 1) if (rem is not None and ndir) else None)
                por_anio[str(anio)] = {"rem": rem, "ndir": ndir, "prom": prom}
            moneda = "-"
            for anio in (anio_n1, anio_n2):
                d = datos.get(anio)
                if d:
                    moneda = d["moneda"]; break
            empresas[emp] = {"sector": sector_de.get(emp, ""),
                             "moneda": moneda, "por_anio": por_anio}

            # comites: usar el anio mas reciente con datos
            cdat = datos.get(anio_n1) or datos.get(anio_n2)
            if cdat:
                comites[emp] = {"n_comites": cdat["comites_n"],
                                "directores_participan": cdat["comites_directores_participan"],
                                "sesiones_total": cdat["comites_sesiones_total"]}

    return {
        "anio_n1": anio_n1, "anio_n2": anio_n2,
        "anios": anios_disponibles,
        "generado": datetime.now().strftime("%d-%m-%Y %H:%M"),
        "sectores": sectores,
        "empresas": empresas,
        "comites": comites,
    }


_HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>__TITLE__</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
<style>
  :root{ --azul:#1F3864; --celeste:#4472C4; --fondo:#D9E1F2; --borde:#BFBFBF; --gris:#808080;}
  *{box-sizing:border-box;}
  body{font-family:'Segoe UI',Arial,sans-serif;margin:0;background:#f4f6fb;color:#1a1a1a;}
  header{background:var(--azul);color:#fff;padding:22px 32px;}
  header h1{margin:0;font-size:20px;font-weight:600;}
  header p{margin:4px 0 0;font-size:13px;opacity:.85;}
  .wrap{max-width:1180px;margin:0 auto;padding:24px 32px 60px;}
  section{background:#fff;border:1px solid var(--borde);border-radius:10px;padding:20px 22px;margin-bottom:26px;box-shadow:0 1px 3px rgba(0,0,0,.05);}
  h2{color:var(--azul);font-size:16px;margin:0 0 14px;border-bottom:2px solid var(--fondo);padding-bottom:8px;}
  table{border-collapse:collapse;width:100%;font-size:13px;}
  th,td{border:1px solid var(--borde);padding:6px 9px;}
  th{background:var(--azul);color:#fff;font-weight:600;text-align:center;}
  td.num{text-align:right;font-variant-numeric:tabular-nums;}
  td.ctr{text-align:center;}
  tr.sector td{background:var(--celeste);color:#fff;font-weight:700;letter-spacing:.3px;}
  tr.empresa:nth-child(even) td{background:#fafbfe;}
  .controls{display:flex;gap:14px;flex-wrap:wrap;margin-bottom:14px;align-items:center;}
  .controls label{font-size:13px;color:var(--azul);font-weight:600;}
  select{padding:5px 9px;border:1px solid var(--borde);border-radius:6px;font-size:13px;background:#fff;}
  .chart-box{position:relative;height:420px;}
  .nota{font-size:12px;color:var(--gris);font-style:italic;margin-top:10px;}
  .pill{display:inline-block;background:var(--fondo);color:var(--azul);border-radius:20px;padding:2px 10px;font-size:12px;font-weight:600;}
</style>
</head>
<body>
<header>
  <h1>Remuneración del Directorio — Estimación bottom-up</h1>
  <p>Generado el __FECHA__ · Valores en miles (000) en la moneda de cada empresa · Estimación aproximada.</p>
</header>
<div class="wrap">

  <section>
    <h2>1 · Tabla principal</h2>
    <div id="tabla-principal"></div>
    <div class="nota">Prom/Director = remuneración estimada de la empresa / n.º de directores, en la moneda de la empresa.</div>
  </section>

  <section>
    <h2>2 · Ranking por remuneración total</h2>
    <div class="controls">
      <label>Año <select id="rk-anio"></select></label>
      <label>Sector <select id="rk-sector"></select></label>
    </div>
    <div class="chart-box"><canvas id="rk-chart"></canvas></div>
  </section>

  <section>
    <h2>3 · Promedio por director (comparativo por sector)</h2>
    <div class="controls">
      <label>Sector <select id="cmp-sector"></select></label>
    </div>
    <div class="chart-box"><canvas id="cmp-chart"></canvas></div>
  </section>

  <section>
    <h2>4 · Evolución histórica</h2>
    <div id="evo-nota"></div>
    <div class="chart-box"><canvas id="evo-chart"></canvas></div>
  </section>

  <section>
    <h2>5 · Comités</h2>
    <div id="tabla-comites"></div>
  </section>

</div>
<script>
const DATA = __DATA__;
const PALETA = ["#1F3864","#4472C4","#8FAADC","#2E75B6","#C55A11","#548235",
                "#7030A0","#BF9000","#D9534F","#5BC0DE","#00876C","#A0522D"];
const fmt = v => (v===null||v===undefined) ? "-" : v.toLocaleString("es-PE");

function empresasOrden(){ // respeta el orden de SECTORES
  let out=[]; DATA.sectores.forEach(s=>s.empresas.forEach(e=>{ if(DATA.empresas[e]) out.push(e); })); return out;
}
function color(i){ return PALETA[i % PALETA.length]; }

/* ---- Seccion 1: tabla principal ---- */
(function(){
  const n1=DATA.anio_n1, n2=DATA.anio_n2;
  let h = "<table><thead><tr><th>Empresa</th><th>Moneda</th>"
        + "<th>Remun "+n1+"</th><th>Remun "+n2+"</th>"
        + "<th>N° Dir "+n1+"</th><th>N° Dir "+n2+"</th>"
        + "<th>Prom/Dir "+n1+"</th><th>Prom/Dir "+n2+"</th></tr></thead><tbody>";
  DATA.sectores.forEach(sec=>{
    const emps = sec.empresas.filter(e=>DATA.empresas[e]);
    if(!emps.length) return;
    h += "<tr class='sector'><td colspan='8'>"+sec.nombre+"</td></tr>";
    emps.forEach(e=>{
      const E=DATA.empresas[e], a1=E.por_anio[n1]||{}, a2=E.por_anio[n2]||{};
      h += "<tr class='empresa'><td>"+e+"</td><td class='ctr'>"+E.moneda+"</td>"
         + "<td class='num'>"+fmt(a1.rem)+"</td><td class='num'>"+fmt(a2.rem)+"</td>"
         + "<td class='ctr'>"+fmt(a1.ndir)+"</td><td class='ctr'>"+fmt(a2.ndir)+"</td>"
         + "<td class='num'>"+fmt(a1.prom)+"</td><td class='num'>"+fmt(a2.prom)+"</td></tr>";
    });
  });
  h += "</tbody></table>";
  document.getElementById("tabla-principal").innerHTML = h;
})();

/* ---- Seccion 2: ranking ---- */
(function(){
  const selA=document.getElementById("rk-anio"), selS=document.getElementById("rk-sector");
  [DATA.anio_n1, DATA.anio_n2].forEach(a=>selA.add(new Option(a,a)));
  selS.add(new Option("Todos","__all__"));
  DATA.sectores.forEach(s=>selS.add(new Option(s.nombre,s.nombre)));
  let chart=null;
  function render(){
    const anio=selA.value, sector=selS.value;
    let emps = empresasOrden().filter(e=> sector==="__all__" || DATA.empresas[e].sector===sector);
    let rows = emps.map(e=>({e, v:(DATA.empresas[e].por_anio[anio]||{}).rem, m:DATA.empresas[e].moneda}))
                   .filter(r=> r.v!==null && r.v!==undefined)
                   .sort((a,b)=>b.v-a.v);
    const labels=rows.map(r=>r.e), vals=rows.map(r=>r.v);
    if(chart) chart.destroy();
    chart=new Chart(document.getElementById("rk-chart"),{
      type:"bar",
      data:{labels,datasets:[{label:"Remuneración (000) "+anio,data:vals,
            backgroundColor:labels.map((_,i)=>color(i))}]},
      options:{indexAxis:"y",responsive:true,maintainAspectRatio:false,
        plugins:{legend:{display:false},
          tooltip:{callbacks:{label:c=>fmt(c.parsed.x)+" mil "+(rows[c.dataIndex].m||"")}}},
        scales:{x:{ticks:{callback:v=>fmt(v)}}}}
    });
  }
  selA.onchange=render; selS.onchange=render; render();
})();

/* ---- Seccion 3: comparativo prom/director por sector ---- */
(function(){
  const selS=document.getElementById("cmp-sector");
  DATA.sectores.forEach((s,i)=>selS.add(new Option(s.nombre,s.nombre)));
  let chart=null;
  function render(){
    const sector=selS.value, n1=DATA.anio_n1, n2=DATA.anio_n2;
    const emps=(DATA.sectores.find(s=>s.nombre===sector)||{empresas:[]}).empresas.filter(e=>DATA.empresas[e]);
    const d1=emps.map(e=>(DATA.empresas[e].por_anio[n1]||{}).prom);
    const d2=emps.map(e=>(DATA.empresas[e].por_anio[n2]||{}).prom);
    if(chart) chart.destroy();
    chart=new Chart(document.getElementById("cmp-chart"),{
      type:"bar",
      data:{labels:emps,datasets:[
        {label:"Prom/Dir "+n1,data:d1,backgroundColor:"#1F3864"},
        {label:"Prom/Dir "+n2,data:d2,backgroundColor:"#8FAADC"}]},
      options:{responsive:true,maintainAspectRatio:false,
        plugins:{tooltip:{callbacks:{label:c=>c.dataset.label+": "+fmt(c.parsed.y)+" mil"}}},
        scales:{y:{ticks:{callback:v=>fmt(v)}}}}
    });
  }
  selS.onchange=render; render();
})();

/* ---- Seccion 4: evolucion historica ---- */
(function(){
  const anios=DATA.anios.slice().sort((a,b)=>a-b);
  if(anios.length<=2){
    document.getElementById("evo-nota").innerHTML="<span class='pill'>solo 2 años disponibles</span>";
  }
  const emps=empresasOrden();
  const datasets=emps.map((e,i)=>({
    label:e,
    data:anios.map(a=>(DATA.empresas[e].por_anio[a]||{}).rem),
    borderColor:color(i),backgroundColor:color(i),tension:.25,spanGaps:true,fill:false
  }));
  new Chart(document.getElementById("evo-chart"),{
    type:"line",
    data:{labels:anios,datasets},
    options:{responsive:true,maintainAspectRatio:false,
      plugins:{tooltip:{callbacks:{label:c=>c.dataset.label+": "+fmt(c.parsed.y)+" mil"}}},
      scales:{y:{ticks:{callback:v=>fmt(v)}}}}
  });
})();

/* ---- Seccion 5: tabla de comites ---- */
(function(){
  let h="<table><thead><tr><th>Empresa</th><th>N° comités</th>"
       +"<th>Directores en ≥1 comité</th><th>Sesiones de comité (total)</th></tr></thead><tbody>";
  DATA.sectores.forEach(sec=>{
    const emps=sec.empresas.filter(e=>DATA.comites[e]);
    if(!emps.length) return;
    h+="<tr class='sector'><td colspan='4'>"+sec.nombre+"</td></tr>";
    emps.forEach(e=>{
      const C=DATA.comites[e];
      h+="<tr class='empresa'><td>"+e+"</td><td class='ctr'>"+fmt(C.n_comites)+"</td>"
        +"<td class='ctr'>"+fmt(C.directores_participan)+"</td>"
        +"<td class='ctr'>"+fmt(C.sesiones_total)+"</td></tr>";
    });
  });
  h+="</tbody></table>";
  document.getElementById("tabla-comites").innerHTML=h;
})();
</script>
</body>
</html>
"""


def escribir_dashboard(ruta_salida, payload):
    html = (_HTML_TEMPLATE
            .replace("__TITLE__", "Remuneración del Directorio — Dashboard")
            .replace("__FECHA__", payload["generado"])
            .replace("__DATA__", json.dumps(payload, ensure_ascii=False)))
    Path(ruta_salida).write_text(html, encoding="utf-8")
    return ruta_salida


# %% ===========================================================================
#    MAIN
# ==============================================================================

def main():
    print("=" * 70)
    print("ESTIMACION BOTTOM-UP DE REMUNERACION DEL DIRECTORIO")
    print("=" * 70)

    hoy = datetime.now()
    anio_n1 = hoy.year - 1          # ano anterior a la fecha de ejecucion
    anio_n2 = hoy.year - 2          # dos anos antes

    # --- FUENTE 1 ---
    print(f"\nLeyendo consolidado:\n  {RUTA_CONSOLIDADO}")
    registros, fuera_cons, hubo_none = leer_consolidado(RUTA_CONSOLIDADO)

    # --- FUENTE 2 ---
    ruta_dietas = hallar_archivo_dietas(RUTA_BASE, PATRON_DIETAS)
    if ruta_dietas is None:
        raise FileNotFoundError(
            f"No se encontro ningun archivo de dietas con patron '{PATRON_DIETAS}' "
            f"en {RUTA_BASE}")
    print(f"\nArchivo de dietas (mas reciente):\n  {ruta_dietas}")
    tarifas, fuera_dietas = leer_tarifas(ruta_dietas)

    # --- Calculo ---
    resultados, anios_disponibles = calcular(registros, tarifas)

    # --- OUTPUT 1 ---
    print(f"\nEscribiendo hoja en el maestro:\n  {RUTA_MAESTRO}")
    nombre_hoja = escribir_hoja_excel(RUTA_MAESTRO, resultados, anio_n1, anio_n2)

    # --- OUTPUT 2 ---
    payload = _construir_payload(resultados, anios_disponibles, anio_n1, anio_n2)
    ruta_dashboard = RUTA_MAESTRO.parent / f"Remuneracion_Directorio_Dashboard_{hoy.strftime('%Y%m%d')}.html"
    escribir_dashboard(ruta_dashboard, payload)

    # --- Resumen en consola ---
    empresas_proc = sorted({r["empresa"] for r in registros})

    # advertencias de equivalencia: deduplicadas por nombre normalizado y
    # limpiando etiquetas tipo "[SIN MATCH PORTAFOLIO]"
    fuera_limpio = {}
    for nombre in (fuera_cons | fuera_dietas):
        display = re.sub(r"\s*\[?\s*sin\s+match\s+portafolio\s*\]?", "",
                         str(nombre), flags=re.IGNORECASE).strip()
        fuera_limpio.setdefault(normalizar(display), display)
    fuera_total = sorted(fuera_limpio.values())

    print("\n" + "-" * 70)
    print("RESUMEN")
    print("-" * 70)
    print(f"Empresas procesadas: {len(empresas_proc)}")
    print(f"  {', '.join(empresas_proc)}")
    print(f"Años analizados: {anio_n1}, {anio_n2}")
    if anios_disponibles:
        print(f"  (años disponibles en el consolidado: "
              f"{', '.join(str(a) for a in anios_disponibles)})")
    sin_tarifa = sorted(e for e in empresas_proc if e not in tarifas)
    if sin_tarifa:
        print(f"Empresas sin tarifa 'Por sesión' (remuneración = '-'): "
              f"{', '.join(sin_tarifa)}")
    print(f"Advertencias de equivalencia (sin match canónico): "
          f"{', '.join(fuera_total) if fuera_total else 'ninguna'}")
    print(f"Registro con ejercicio None ignorado: {'sí' if hubo_none else 'no'}")
    print(f"Hoja Excel generada: {nombre_hoja}")
    print(f"Dashboard HTML: {ruta_dashboard}")
    print("-" * 70)


if __name__ == "__main__":
    main()