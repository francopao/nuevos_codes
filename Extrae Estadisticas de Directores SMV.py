

# -*- coding: utf-8 -*-
"""
Extracción MASIVA y CONSOLIDADA de retribución de Directorio / Alta Gerencia /
Comités especiales desde los PDF 'SECCIONES B Y C' (Reporte de Cumplimiento del
Código BGC) descargados por 'descarga_memorias_smv.py'.

Recorre TODAS las subcarpetas de empresa dentro de RUTA_BASE, procesa cada PDF,
deduplica los documentos que la SMV subió dos veces, y genera UN solo Excel
consolidado con UNA HOJA POR EMPRESA (nombre de hoja = nombre reducido), donde
cada hoja apila, por ejercicio, las secciones relevantes:

    1) Resumen del Directorio (nº directores, independientes, tercio, sesiones,
       esquema y criterios de retribución, si aplica el modelo por sesión).
    2) Retribución como % de ingresos brutos: Directores (indep / no indep) y
       Alta Gerencia (fija / variable)  <-- tabla de la hoja ~46 (Pregunta III.19 a).
    3) Comités especiales (denominación, fecha, sesiones, % indep, participa JGA).
    4) MATRIZ DE PARTICIPACIÓN: filas = directores; columnas = Directorio + cada
       comité con su nº de sesiones. Es el insumo directo del bottom-up:
           sueldo = tarifa_sesion_dir * sesiones_directorio
                  + Σ (tarifa_sesion_comité * sesiones del comité donde participa)

Cada dato lleva una columna 'Pág.' con la hoja del PDF de donde salió, para que
puedas verificarlo rápido.

ADVERTENCIA (no es bug, es el dato): el reporte solo informa la retribución como
% de ingresos brutos; NO da el monto por sesión. Y el criterio de pago varía por
empresa: la fórmula por sesión SOLO aplica donde el criterio fijo marque
"Por sesión de Directorio" (columna 'aplica_modelo_por_sesion').

Sin xlwings/win32com. Comentarios en español. Estructura por celdas (# %%) Spyder.
Dependencias:  pip install pdfplumber openpyxl pandas
"""

# %% ===========================  IMPORTS  ====================================
import os
import re
import glob
import unicodedata
from collections import defaultdict
from difflib import SequenceMatcher

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# %% ===========================  PARÁMETROS  =================================

RUTA_BASE = os.environ.get(
    "RUTA_BASE_OVERRIDE",
    r"C:\Users\usuario\OneDrive\Desktop\AFP INTEGRA\Docs privados\Otros\Remuneracion Directorio - Bottom Up\Remuneraciones",
)
SALIDA_XLSX = os.path.join(RUTA_BASE, "_consolidado_remuneracion_directorio.xlsx")

PATRON_ARCHIVO = "SECCIONES_B_Y_C"   # patrón en el nombre de archivo
UMBRAL_NOMBRE = 0.82                  # similitud mínima para emparejar nombres


# %% ======================  UTILIDADES DE TEXTO  ============================

def norm(s):
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip()


def nlow(s):
    return norm(s).lower()


FECHA = re.compile(r"\b(\d{2}/\d{2}/\d{4})\b")
TITULO_COMITE = re.compile(r"^COMIT[EE]\s*(\d+)$")
CARGOS_COMITE = ("presidente", "vicepresidente", "titular", "suplente",
                 "alterno", "secretario", "miembro")


def primer_numero(celdas):
    for c in celdas:
        m = re.search(r"-?\d+(?:\.\d+)?", norm(c))
        if m:
            return m.group(0)
    return None


def numeros_en(celdas):
    out = []
    for c in celdas:
        m = re.search(r"-?\d+(?:\.\d+)?", norm(c))
        if m:
            out.append(m.group(0))
    return out


def fila_si_no(celdas):
    idx_si = idx_no = idx_x = None
    for i, c in enumerate(celdas):
        n = nlow(c)
        if n in ("si", "sí") and idx_si is None:
            idx_si = i
        if n == "no" and idx_no is None:
            idx_no = i
        if n == "x" and idx_x is None:
            idx_x = i
    if idx_x is None or (idx_si is None and idx_no is None):
        return None
    if idx_si is not None and idx_no is not None:
        return "Sí" if abs(idx_x - idx_si) <= abs(idx_x - idx_no) else "No"
    return "Sí" if idx_si is not None else "No"


def parece_nombre(c):
    cl = nlow(c)
    if not c or FECHA.search(c) or len(c) < 5:
        return False
    if cl in CARGOS_COMITE:
        return False
    if any(k in cl for k in ["cargo", "fecha", "inicio", "termino", "facultades",
                             "ley general", "brindara", "comite", "sesiones",
                             "denominacion", "principales", "directores independientes",
                             "gerente general"]):
        return False
    return sum(1 for w in c.split() if w[:1].isupper()) >= 2


def similar(a, b):
    return SequenceMatcher(None, nlow(a), nlow(b)).ratio()


# %% ======================  EXTRACCIÓN POR PDF  =============================

def _flat_rows(pdf):
    """Filas de todas las tablas, cada una como (pagina_1based, [celdas norm])."""
    out = []
    for pi, pg in enumerate(pdf.pages):
        for tb in pg.extract_tables():
            for row in tb:
                out.append((pi + 1, [norm(c) for c in row]))
    return out


def _pagina_de(textos_pag, ancla):
    """Devuelve la 1ª página (1-based) cuyo texto contiene 'ancla' (normalizado)."""
    a = nlow(ancla)
    for i, t in enumerate(textos_pag):
        if a in nlow(t):
            return i + 1
    return None


def extrae_pdf(ruta):
    with pdfplumber.open(ruta) as pdf:
        rows = _flat_rows(pdf)
        textos_pag = [(pg.extract_text() or "") for pg in pdf.pages]
    texto = "\n".join(textos_pag)
    p1 = textos_pag[0] if textos_pag else ""

    d = {"archivo": os.path.basename(ruta), "ruta": ruta, "pag": {}}

    # --- Empresa / ejercicio (pág. 1) ---
    emp = None
    lineas = p1.split("\n")
    for i, l in enumerate(lineas):
        if nlow(l).startswith("denominacion:") and i + 1 < len(lineas):
            emp = norm(lineas[i + 1])
            break
    d["empresa"] = emp
    mej = re.search(r"Ejercicio:\s*(\d{4})", p1)
    d["ejercicio"] = int(mej.group(1)) if mej else None

    # --- Independientes (Principio 19) ---
    m_tot = re.search(r"De los\s+(\d+)\s+directores que conforman", texto)
    m_ind = re.search(r"(\d+)\s+son directores independientes", texto)
    d["n_directores_total"] = int(m_tot.group(1)) if m_tot else None
    d["n_independientes"] = int(m_ind.group(1)) if m_ind else None
    d["pag"]["independientes"] = _pagina_de(textos_pag, "son directores independientes")

    # ¿cumple el tercio? (arrastra encabezado Sí/No de la fila previa)
    d["cumple_tercio"] = None
    hdr_si = hdr_no = None
    for _, r in rows:
        c_si = next((i for i, c in enumerate(r) if nlow(c) in ("si", "sí")), None)
        c_no = next((i for i, c in enumerate(r) if nlow(c) == "no"), None)
        if c_si is not None and c_no is not None:
            hdr_si, hdr_no = c_si, c_no
        if "tercio del directorio" in nlow(" ".join(r)):
            sn = fila_si_no(r)
            if sn is None:
                idx_x = next((i for i, c in enumerate(r) if nlow(c) == "x"), None)
                if idx_x is not None and hdr_si is not None and hdr_no is not None:
                    sn = "Sí" if abs(idx_x - hdr_si) <= abs(idx_x - hdr_no) else "No"
            d["cumple_tercio"] = sn
            break

    # --- Retribución directores: esquema (f) y criterios (g) ---
    OPC = {
        "esquema": ["Fijo", "Variable", "Mixto (fijo + variable)"],
        "criterio_fijo": ["Por sesion de Directorio", "Por sesion de Comite",
                          "Por mes", "Por ano"],
        "criterio_variable": ["Por resultados del ejercicio",
                              "Por cumplimiento de objetivos"],
    }
    marcados = {k: [] for k in OPC}
    pag_esquema = None
    for pg, r in rows:
        if not any(nlow(c) == "x" for c in r):
            continue
        for grp, ops in OPC.items():
            for op in ops:
                if any(nlow(c) == nlow(op) for c in r):
                    marcados[grp].append(op)
                    if grp != "esquema":
                        pag_esquema = pg
    d["esquema_retribucion"] = ", ".join(dict.fromkeys(marcados["esquema"])) or None
    d["criterio_fijo"] = ", ".join(dict.fromkeys(marcados["criterio_fijo"])) or None
    d["criterio_variable"] = ", ".join(dict.fromkeys(marcados["criterio_variable"])) or None
    d["aplica_modelo_por_sesion"] = bool(
        re.search(r"por sesion de directorio", nlow(d["criterio_fijo"] or "")))
    d["paga_sesion_comite"] = bool(
        re.search(r"por sesion de comite", nlow(d["criterio_fijo"] or "")))
    d["pag"]["esquema"] = pag_esquema

    # --- Retribución directores como % de ingresos brutos (e) ---
    d["retrib_pct_no_indep"] = None
    d["retrib_pct_indep"] = None
    en_e = False
    for pg, r in rows:
        j = nlow(" ".join(r))
        if "retribuciones y de las bonificaciones" in j:
            en_e = True
            d["pag"]["retrib_dir"] = pg
            continue
        if en_e:
            if "directores (sin incluir" in j:
                d["retrib_pct_no_indep"] = primer_numero(
                    [c for c in r if not parece_nombre(c)])
            elif j.startswith("directores independientes"):
                d["retrib_pct_indep"] = primer_numero(r)
            elif "esquema de retribucion" in j or "f. precise" in j:
                en_e = False

    # --- Retribución Alta Gerencia (Pregunta III.19 a): Fija / Variable ---
    d["ger_fija_pct"] = None
    d["ger_variable_pct"] = None
    for pg, r in rows:
        if any("gerente general y plana gerencial" == nlow(c) for c in r):
            nums = numeros_en(r)
            if len(nums) >= 2:
                d["ger_fija_pct"], d["ger_variable_pct"] = nums[0], nums[1]
            elif len(nums) == 1:
                d["ger_fija_pct"] = nums[0]
            d["pag"]["retrib_gerencia"] = pg
            break

    # --- Sesiones del Directorio + asistencia por director (Principio 20) ---
    m_ses = re.search(r"N[uú]mero de sesiones realizadas\s+(\d+)", texto)
    d["n_sesiones_directorio"] = int(m_ses.group(1)) if m_ses else None

    asistencia = []
    en_tabla = False
    for pg, r in rows:
        j = nlow(" ".join(r))
        if "nombres y apellidos" in j and "convocadas" in j:
            en_tabla = True
            d["pag"]["sesiones"] = pg
            continue
        if en_tabla:
            nombre = next((c for c in r if parece_nombre(c)), None)
            nums = [c for c in r if re.fullmatch(r"\d+", norm(c))]
            if nombre and len(nums) >= 2:
                asistencia.append({"director": nombre,
                                   "sesiones_convocadas": int(nums[0]),
                                   "sesiones_asistio": int(nums[1])})
            elif r and not nombre and not nums and any(len(norm(c)) > 40 for c in r):
                en_tabla = False
    d["asistencia_directorio"] = asistencia
    if d["n_directores_total"] is None and asistencia:
        d["n_directores_total"] = len(asistencia)

    # --- Bloques COMITÉ N (Principio 21 'b.') ---
    comites = []
    cur = None
    for pg, r in rows:
        es_titulo = any(TITULO_COMITE.fullmatch(norm(c).upper()) for c in r)
        if es_titulo:
            if cur:
                comites.append(cur)
            idx = next(TITULO_COMITE.fullmatch(norm(c).upper()).group(1)
                       for c in r if TITULO_COMITE.fullmatch(norm(c).upper()))
            cur = {"comite_idx": int(idx), "denominacion": None, "fecha_creacion": None,
                   "n_sesiones": None, "pct_indep": None, "participa_jga": None,
                   "pagina": pg, "miembros": []}
            continue
        if cur is None:
            continue
        j = nlow(" ".join(r))
        if "denominacion del comite" in j and cur["denominacion"] is None:
            v = [c for c in r if c and "denominacion" not in nlow(c)]
            cur["denominacion"] = v[0] if v else None
        elif "fecha de creacion" in j and cur["fecha_creacion"] is None:
            mf = next((FECHA.search(c) for c in r if FECHA.search(c)), None)
            if mf:
                cur["fecha_creacion"] = mf.group(1)
        elif "numero de sesiones realizadas" in j:
            cur["n_sesiones"] = primer_numero(
                [c for c in r if "numero de sesiones" not in nlow(c)])
        elif "directores independientes respec" in j:
            cur["pct_indep"] = primer_numero(
                [c for c in r if "independientes" not in nlow(c)])
        elif "participa en la jga" in j:
            cur["participa_jga"] = fila_si_no(r)
        else:
            nombre = next((c for c in r if parece_nombre(c)), None)
            if nombre and any(FECHA.search(c) for c in r):
                cargo = next((c for c in r if nlow(c) in CARGOS_COMITE), None)
                cur["miembros"].append({"nombre": nombre, "cargo_comite": cargo})
    if cur:
        comites.append(cur)
    d["comites"] = [c for c in comites if c["denominacion"]]
    d["n_comites_especiales"] = len(d["comites"])

    # emparejar nombres de integrantes con el canónico del directorio
    d["incidencias_nombre"] = _canoniza_miembros(d)
    return d


def _canoniza_miembros(d):
    canon = [a["director"] for a in d.get("asistencia_directorio", [])]
    inc = []
    for com in d.get("comites", []):
        for m in com["miembros"]:
            if not canon:
                m["director_canonico"] = m["nombre"]
                continue
            mejor = max(canon, key=lambda x: similar(x, m["nombre"]))
            if similar(mejor, m["nombre"]) >= UMBRAL_NOMBRE:
                m["director_canonico"] = mejor
            else:
                m["director_canonico"] = m["nombre"]
                inc.append({"empresa": d.get("empresa"), "ejercicio": d.get("ejercicio"),
                            "comite": com["denominacion"], "nombre_pdf": m["nombre"],
                            "mejor_candidato": mejor,
                            "score": round(similar(mejor, m["nombre"]), 2)})
    return inc


# %% ======================  ESTILOS / HELPERS EXCEL  =======================

F_TITULO = Font(bold=True, size=13, color="1F3864")
F_EJERC = Font(bold=True, size=11, color="FFFFFF")
F_SECCION = Font(bold=True, size=10, color="1F3864")
F_HEADER = Font(bold=True, size=9, color="FFFFFF")
F_NOTA = Font(italic=True, size=8, color="808080")
FILL_EJERC = PatternFill("solid", fgColor="1F3864")
FILL_HEADER = PatternFill("solid", fgColor="4472C4")
FILL_SECCION = PatternFill("solid", fgColor="D9E1F2")
FILL_APLICA = PatternFill("solid", fgColor="C6EFCE")
FILL_NOAPLICA = PatternFill("solid", fgColor="FFC7CE")
THIN = Side(style="thin", color="BFBFBF")
BORDE = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")


def _corto_comite(denom):
    """Nombre corto de comité para encabezado de la matriz."""
    s = norm(denom)
    s = re.sub(r"^Comite de\s+", "", s, flags=re.I)
    return (s[:22] + "…") if len(s) > 23 else s


def _alias_empresa(emp, usados):
    """Nombre de hoja reducido y único (<=31 chars)."""
    e = norm(emp or "EMPRESA").upper()
    for suf in [" SOCIEDAD ANONIMA ABIERTA", " S.A.A.", " S.A.A", " S.A.C.",
                " S.A.", " S.A", " CORP.", " CORP", " LTD.", " LTD",
                " INC.", " INC", " PERU"]:
        e = e.replace(suf, "")
    e = re.sub(r"[\[\]\:\*\?\/\\]", "", e)
    e = re.sub(r"\s*-\s*.*$", "", e).strip()  # quita la parte tras un guion
    base = (e or "EMPRESA")[:28]
    nombre, k = base, 2
    while nombre in usados:
        nombre = f"{base[:28]}_{k}"
        k += 1
    usados.add(nombre)
    return nombre


def _bloque_tabla(ws, fila, titulo, headers, datos, anchos=None):
    """Escribe un bloque: título de sección + cabecera + filas. Devuelve la
    siguiente fila libre."""
    ncol = len(headers)
    ws.cell(fila, 1, titulo).font = F_SECCION
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=ncol)
    for c in range(1, ncol + 1):
        ws.cell(fila, c).fill = FILL_SECCION
    fila += 1
    for j, h in enumerate(headers, start=1):
        cel = ws.cell(fila, j, h)
        cel.font = F_HEADER
        cel.fill = FILL_HEADER
        cel.alignment = WRAP
        cel.border = BORDE
    fila += 1
    for row in datos:
        for j, v in enumerate(row, start=1):
            cel = ws.cell(fila, j, v)
            cel.border = BORDE
            cel.alignment = CENTER if j > 1 else WRAP
        fila += 1
    return fila + 1  # deja una fila en blanco


# %% ======================  ESCRITURA POR EMPRESA  =========================

def escribir_hoja_empresa(ws, registros):
    """registros: lista de dicts de UNA empresa (uno por ejercicio)."""
    regs = sorted(registros, key=lambda d: (d.get("ejercicio") or 0), reverse=True)
    emp = next((r.get("empresa") for r in regs if r.get("empresa")), "EMPRESA")

    ws.cell(1, 1, emp).font = F_TITULO
    ws.freeze_panes = "A3"
    fila = 3

    for d in regs:
        ej = d.get("ejercicio")
        ws.cell(fila, 1, f"EJERCICIO {ej}   ·   archivo: {d.get('archivo')}").font = F_EJERC
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=9)
        for c in range(1, 10):
            ws.cell(fila, c).fill = FILL_EJERC
        fila += 2

        pg = d.get("pag", {})
        # --- 1) Resumen del Directorio ---
        fila = _bloque_tabla(
            ws, fila, "1. Resumen del Directorio",
            ["N° Directores", "N° Independientes", "Cumple tercio",
             "Sesiones Directorio", "Esquema retribución", "Criterio fijo",
             "Criterio variable", "¿Aplica modelo por sesión?", "Pág."],
            [[d.get("n_directores_total"), d.get("n_independientes"),
              d.get("cumple_tercio"), d.get("n_sesiones_directorio"),
              d.get("esquema_retribucion"), d.get("criterio_fijo"),
              d.get("criterio_variable"),
              "SÍ" if d.get("aplica_modelo_por_sesion") else "NO",
              ", ".join(str(pg.get(k)) for k in ("independientes", "esquema", "sesiones")
                        if pg.get(k))]])
        # colorear la celda "aplica"
        cel_aplica = ws.cell(fila - 3, 8)
        cel_aplica.fill = FILL_APLICA if d.get("aplica_modelo_por_sesion") else FILL_NOAPLICA

        # --- 2) Retribución (% de ingresos brutos) ---
        fila = _bloque_tabla(
            ws, fila, "2. Retribución (% de ingresos brutos)",
            ["Directores no indep.", "Directores indep.",
             "Alta Gerencia – Fija", "Alta Gerencia – Variable",
             "Pág. directores", "Pág. gerencia"],
            [[d.get("retrib_pct_no_indep"), d.get("retrib_pct_indep"),
              d.get("ger_fija_pct"), d.get("ger_variable_pct"),
              pg.get("retrib_dir"), pg.get("retrib_gerencia")]])

        # --- 3) Comités especiales ---
        comites = d.get("comites", [])
        datos_com = [[c["denominacion"], c["fecha_creacion"], c["n_sesiones"],
                      c["pct_indep"], c["participa_jga"], len(c["miembros"]),
                      c["pagina"]] for c in comites]
        fila = _bloque_tabla(
            ws, fila, f"3. Comités especiales ({len(comites)})",
            ["Comité", "Fecha creación", "Sesiones", "% Indep.",
             "Participa JGA", "N° miembros", "Pág."],
            datos_com or [["(sin comités detectados)", "", "", "", "", "", ""]])

        # --- 4) Matriz de participación (bottom-up) ---
        asist = d.get("asistencia_directorio", [])
        asist_map = {a["director"]: a for a in asist}
        # miembros canónicos por comité
        sets_com = [{m.get("director_canonico", m["nombre"]) for m in c["miembros"]}
                    for c in comites]
        head = (["Director", "Ses. Dir. asistió", "Ses. Dir. convoc."]
                + [f"{_corto_comite(c['denominacion'])} ({c['n_sesiones']} ses)"
                   for c in comites]
                + ["Σ ses. comités"])
        filas_matriz = []
        # universo de directores = asistencia ∪ miembros de comités
        universo = [a["director"] for a in asist]
        for s in sets_com:
            for nom in s:
                if nom not in universo:
                    universo.append(nom)
        for nom in universo:
            a = asist_map.get(nom)
            row = [nom,
                   a["sesiones_asistio"] if a else "",
                   a["sesiones_convocadas"] if a else ""]
            total = 0
            for c, s in zip(comites, sets_com):
                if nom in s:
                    n = int(c["n_sesiones"]) if str(c["n_sesiones"]).isdigit() else None
                    row.append(n if n is not None else "✓")
                    total += (n or 0)
                else:
                    row.append("")
            row.append(total)
            filas_matriz.append(row)
        fila = _bloque_tabla(
            ws, fila,
            "4. Matriz de participación  (sesiones por órgano · insumo bottom-up)",
            head, filas_matriz or [["(sin datos)"] + [""] * (len(head) - 1)])

        ws.cell(fila, 1, "Nota: el reporte no informa monto por sesión ni asistencia "
                         "por miembro a cada comité; multiplique estas cantidades por "
                         "su tarifa por sesión.").font = F_NOTA
        fila += 3  # separación entre ejercicios

    # anchos de columna
    ws.column_dimensions["A"].width = 34
    for col in range(2, 14):
        ws.column_dimensions[get_column_letter(col)].width = 16


# %% ======================  EJECUCIÓN PRINCIPAL  ===========================

def localizar_pdfs(base):
    todos = glob.glob(os.path.join(base, "**", "*.pdf"), recursive=True)
    return sorted(p for p in todos
                  if PATRON_ARCHIVO.lower() in os.path.basename(p).lower())


def main():
    pdfs = localizar_pdfs(RUTA_BASE)
    print(f"PDFs candidatos encontrados: {len(pdfs)}")

    registros, incidencias = [], []
    vistos = {}                      # (empresa_norm, ejercicio) -> archivo (dedupe)
    por_empresa = defaultdict(list)  # empresa_norm -> [registros]

    for ruta in pdfs:
        try:
            d = extrae_pdf(ruta)
        except Exception as e:
            incidencias.append({"tipo": "error_lectura",
                                "archivo": os.path.basename(ruta), "detalle": str(e)})
            print(f"  [!] error en {os.path.basename(ruta)}: {e}")
            continue

        clave = (nlow(d.get("empresa") or os.path.dirname(ruta)), d.get("ejercicio"))
        if clave in vistos and d.get("empresa"):
            incidencias.append({"tipo": "duplicado", "empresa": d.get("empresa"),
                                "ejercicio": d.get("ejercicio"),
                                "archivo_omitido": d.get("archivo"),
                                "archivo_conservado": vistos[clave]})
            print(f"  · duplicado omitido: {d.get('archivo')}")
            continue
        vistos[clave] = d.get("archivo")

        incidencias.extend(d.pop("incidencias_nombre", []))
        registros.append(d)
        por_empresa[nlow(d.get("empresa") or os.path.dirname(ruta))].append(d)
        print(f"  OK {d.get('empresa')} {d.get('ejercicio')} | "
              f"dir={d.get('n_directores_total')} indep={d.get('n_independientes')} "
              f"comites={d.get('n_comites_especiales')} "
              f"esquema={d.get('esquema_retribucion')} | ger_fija={d.get('ger_fija_pct')}")

    if not registros:
        print("No se extrajo ningún registro. Revisa RUTA_BASE y el patrón.")
        return

    # --- construir el libro ---
    wb = Workbook()
    wb.remove(wb.active)

    # hoja índice
    idx = wb.create_sheet("Índice")
    idx.cell(1, 1, "ÍNDICE DE EMPRESAS").font = F_TITULO
    idx.append([])
    for j, h in enumerate(["Empresa", "Hoja", "Ejercicios", "N° comités (último)"], start=1):
        c = idx.cell(3, j, h); c.font = F_HEADER; c.fill = FILL_HEADER; c.border = BORDE
    idx.freeze_panes = "A4"

    usados = set()
    fila_idx = 4
    for emp_key, regs in sorted(por_empresa.items()):
        emp = next((r.get("empresa") for r in regs if r.get("empresa")), emp_key)
        hoja = _alias_empresa(emp, usados)
        ws = wb.create_sheet(hoja)
        escribir_hoja_empresa(ws, regs)
        ejs = ", ".join(str(e) for e in sorted(
            {r.get("ejercicio") for r in regs if r.get("ejercicio")}, reverse=True))
        ult = max(regs, key=lambda r: (r.get("ejercicio") or 0))
        for j, v in enumerate([emp, hoja, ejs, ult.get("n_comites_especiales")], start=1):
            c = idx.cell(fila_idx, j, v); c.border = BORDE
        fila_idx += 1
    idx.column_dimensions["A"].width = 45
    idx.column_dimensions["B"].width = 22
    idx.column_dimensions["C"].width = 22
    idx.column_dimensions["D"].width = 18

    # hoja incidencias
    inc = wb.create_sheet("Incidencias")
    if incidencias:
        cols = sorted({k for d in incidencias for k in d})
        for j, h in enumerate(cols, start=1):
            c = inc.cell(1, j, h); c.font = F_HEADER; c.fill = FILL_HEADER
        for i, d in enumerate(incidencias, start=2):
            for j, h in enumerate(cols, start=1):
                inc.cell(i, j, d.get(h))
    else:
        inc.cell(1, 1, "Sin incidencias")

    wb.save(SALIDA_XLSX)

    print("\n" + "=" * 70)
    print(f"Empresas (hojas)  : {len(por_empresa)}")
    print(f"Registros (años)  : {len(registros)}")
    print(f"Incidencias       : {len(incidencias)}")
    print(f"Excel generado    : {SALIDA_XLSX}")


# %% ======================  PUNTO DE ENTRADA  ==============================
if __name__ == "__main__":
    main()